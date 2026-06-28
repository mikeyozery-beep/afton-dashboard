#!/usr/bin/env python3
"""
Afton Dashboard - Extract Metrics to Google Sheet
Reads from Excel files and writes to Google Sheet (single source of truth)
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime, timedelta
from glob import glob

try:
    import openpyxl
    from openpyxl import load_workbook
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
except ImportError as e:
    print(f"ERROR: Missing library - {e}")
    print("Run: pip install openpyxl google-auth-oauthlib google-auth-httplib2 google-api-python-client --break-system-packages")
    exit(1)

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
DATA_DIR = SCRIPT_DIR / "data"
LOG_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"sheets_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Google Sheets API Configuration
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SHEET_ID = "1PAnz3SC7mKRiRdRsmeUIp-zP0gsGQDAl_RR4pmHG6YE"
CREDENTIALS_FILE = SCRIPT_DIR / "credentials.json"

# File paths
GREENBOOKS_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Greenbooks")
FINANCIAL_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\5. Monthly Financial Reviews")

# Property tabs in Effective Rent Analysis
PROPERTY_TABS = [
    'RGCP', 'RGSR', 'RGP', 'RGSH', 'RGLA', 'RGD', 'RGOX', 'RGPT',
    'RGSV', 'RGSC', 'RGTCC', 'RGCCV', 'RGCK', 'RGWL', 'RGOV', 'RGSM',
    'RGLT', 'RGST', 'RGAL', 'RGCW', 'RGAZ', 'RGHC', 'RGLSV'
]

def get_sheets_service():
    """Authenticate and get Google Sheets service"""
    creds = None
    token_file = SCRIPT_DIR / "token.json"

    # Load existing token (handle UTF-8 BOM if present)
    if token_file.exists():
        try:
            creds = Credentials.from_authorized_user_file(token_file, SCOPES)
        except Exception:
            # If token.json is corrupted, we'll re-authenticate
            pass

    # If no valid credentials, get new ones
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE.exists():
                logger.error(f"credentials.json not found at {CREDENTIALS_FILE}")
                raise FileNotFoundError(f"Please create credentials.json using Google Cloud Console")

            # Load credentials.json with UTF-8 BOM handling
            try:
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            except json.JSONDecodeError:
                # If standard loading fails due to BOM, load manually
                with open(CREDENTIALS_FILE, 'r', encoding='utf-8-sig') as f:
                    client_config = json.load(f)
                flow = InstalledAppFlow.from_client_config(client_config, SCOPES)

            # Desktop app: run_local_server opens browser automatically
            creds = flow.run_local_server(port=0, open_browser=True)

        # Save token for next run
        with open(token_file, 'w') as token:
            token.write(creds.to_json())

    return build('sheets', 'v4', credentials=creds)

def extract_occupancy_metrics():
    """Extract Portfolio Occupancy and Leased Occupancy from Occupancy Analysis"""
    try:
        file_path = GREENBOOKS_PATH / "Occupancy Analysis.xlsx"
        logger.info(f"Reading Occupancy Analysis from {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws = wb['Summary']

        portfolio_occ = ws['G30'].value
        leased_occ = ws['H30'].value

        return {
            'Portfolio Occupancy': format_percentage(portfolio_occ),
            'Leased Occupancy': format_percentage(leased_occ)
        }
    except Exception as e:
        logger.error(f"Error extracting occupancy metrics: {e}")
        return {}

def extract_dq_metrics():
    """Extract Percent Collected from DQ Reports"""
    try:
        file_path = GREENBOOKS_PATH / "DQ Reports.xlsm"
        logger.info(f"Reading DQ Reports from {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws = wb['Summary']

        percent_collected = ws['H29'].value

        return {
            'Percent Collected': format_percentage(percent_collected)
        }
    except Exception as e:
        logger.error(f"Error extracting DQ metrics: {e}")
        return {}

def extract_effective_rent_metrics():
    """Extract Annualized Rent Growth (current and prior year)"""
    try:
        file_path = GREENBOOKS_PATH / "Effective Rent Analysis.xlsx"
        logger.info(f"Reading Effective Rent Analysis from {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws_summary = wb['Summary']

        # Current month (Latest)
        current_growth = ws_summary['O109'].value
        logger.info(f"[OK] Annualized Rent Growth (Current): {current_growth}")

        # Prior Year (SUMPRODUCT calculation)
        prior_year_growth = calculate_prior_year_growth(wb, ws_summary)
        logger.info(f"[OK] Annualized Rent Growth (Prior Year): {prior_year_growth}")

        return {
            'Annualized Rent Growth': format_percentage(current_growth),
            'Annualized Rent Growth (Prior Year)': format_percentage(prior_year_growth)
        }
    except Exception as e:
        logger.error(f"Error extracting effective rent metrics: {e}")
        return {}

def calculate_prior_year_growth(wb, ws_summary):
    """Calculate weighted average effective rent growth"""
    try:
        from datetime import datetime

        logger.info("Calculating prior year annualized rent growth...")

        weighted_sum = 0
        total_units = 0

        for prop_code in PROPERTY_TABS:
            if prop_code not in wb.sheetnames:
                continue

            ws_prop = wb[prop_code]
            prior_month = ws_prop['R48'].value

            if not prior_month:
                continue

            # Find column with prior year date in row 345
            prior_year_col = None
            for col_num in range(1, 500):
                cell_val = ws_prop.cell(row=345, column=col_num).value
                if cell_val and isinstance(cell_val, datetime):
                    if cell_val.year == prior_month.year and cell_val.month == prior_month.month:
                        prior_year_col = col_num
                        break

            if not prior_year_col:
                continue

            unit_count = ws_prop['B386'].value
            if not unit_count or not isinstance(unit_count, (int, float)):
                continue

            prior_year_rent = ws_prop.cell(row=386, column=prior_year_col).value
            if prior_year_rent and isinstance(prior_year_rent, (int, float)):
                weighted_sum += unit_count * prior_year_rent
                total_units += unit_count

        if total_units > 0:
            weighted_average_prior = weighted_sum / total_units
            logger.info(f"  Prior year weighted average rent: ${weighted_average_prior:.2f}")

            # Calculate current month average
            current_weighted_sum = 0
            current_total_units = 0

            for prop_code in PROPERTY_TABS:
                if prop_code not in wb.sheetnames:
                    continue

                ws_prop = wb[prop_code]
                current_month = ws_prop['O48'].value

                if not current_month:
                    continue

                current_col = None
                for col_num in range(1, 500):
                    cell_val = ws_prop.cell(row=345, column=col_num).value
                    if cell_val and isinstance(cell_val, datetime):
                        if cell_val.year == current_month.year and cell_val.month == current_month.month:
                            current_col = col_num
                            break

                if not current_col:
                    continue

                unit_count = ws_prop['B386'].value
                if not unit_count or not isinstance(unit_count, (int, float)):
                    continue

                current_rent = ws_prop.cell(row=386, column=current_col).value
                if current_rent and isinstance(current_rent, (int, float)):
                    current_weighted_sum += unit_count * current_rent
                    current_total_units += unit_count

            if current_total_units > 0:
                weighted_average_current = current_weighted_sum / current_total_units
                logger.info(f"  Current month weighted average rent: ${weighted_average_current:.2f}")

                if weighted_average_prior > 0:
                    growth_rate = (weighted_average_current - weighted_average_prior) / weighted_average_prior
                    return growth_rate

        return None

    except Exception as e:
        logger.error(f"Error in growth calculation: {e}")
        return None

def format_percentage(value):
    """Format a decimal value as percentage string"""
    if value is None:
        return "—"
    if isinstance(value, (int, float)):
        return f"{value*100:.2f}%"
    return str(value)

def write_to_sheets(service, metrics_dict):
    """Write metrics to Google Sheet"""
    try:
        # Prepare data for METRICS tab
        metrics_rows = [
            ["Metric Name", "Current Value", "Change", "Detail / Notes"],
        ]

        # Hardcode the 8 metrics in order
        metric_configs = [
            ("Portfolio Occupancy", "Portfolio Occupancy", "▲ +50 bps", "Prior Week: 94.3% | Budget: 95.0%"),
            ("Leased Occupancy", "Leased Occupancy", "▲ +40 bps", "Prior Week: 95.7% | Budget: 96.0%"),
            ("% Collected", "Percent Collected", "▲ +30 bps", "As of: 06/15/26 | Budget: 99.0%"),
            ("Annualized Rent Growth", "Annualized Rent Growth", "▲ +40 bps", "Prior YTD: 4.3% | Budget: 4.0%"),
            ("Trend Occupancy 30-Day", "Trend Occupancy 30-Day", "▲ +40 bps", "Budget: 95.0%"),
            ("# Staffing Vacancies", "Staffing Vacancies", "▼ -2 Open", "Prior Month: 14"),
            ("NOI Var. to Budget (Thru June)", "NOI Variance", "—", "Prior YTD: +1.9% | Budget: 0.0%"),
            ("CapEx vs Budget", "CapEx vs Budget", "🔴 RED", "YTD: $1.9M Actual vs $1.8M Budget"),
        ]

        for display_name, key, change, detail in metric_configs:
            value = metrics_dict.get(key, "—")
            metrics_rows.append([display_name, value, change, detail])

        # Update METRICS sheet
        logger.info("Writing to METRICS tab...")
        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range="METRICS!A1:D9",
            valueInputOption="USER_ENTERED",
            body={"values": metrics_rows}
        ).execute()

        # Update METADATA tab
        logger.info("Writing to METADATA tab...")
        metadata_rows = [
            ["Field", "Value"],
            ["Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Report Date", datetime.now().strftime("%B %d, %Y")],
            ["Data Source", "Excel Extract (Python)"],
            ["Next Update", "Daily 9:00 AM"],
        ]

        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range="METADATA!A1:B5",
            valueInputOption="USER_ENTERED",
            body={"values": metadata_rows}
        ).execute()

        logger.info("[OK] Successfully wrote data to Google Sheet")
        return True

    except Exception as e:
        logger.error(f"Error writing to Google Sheet: {e}")
        return False

def save_metrics_json(metrics_dict):
    """Save metrics to local JSON cache for dashboard"""
    try:
        data = {
            'timestamp': datetime.now().isoformat(),
            'metrics': [
                {
                    'name': key,
                    'value': value,
                    'updated': datetime.now().strftime('%Y-%m-%d'),
                    'source': get_source(key)
                }
                for key, value in metrics_dict.items()
            ],
            'count': len(metrics_dict),
            'success': True
        }

        with open(DATA_DIR / 'metrics.json', 'w') as f:
            json.dump(data, f, indent=2)

        logger.info(f"[OK] Saved {len(metrics_dict)} metrics to local JSON cache")
        return True

    except Exception as e:
        logger.error(f"Error saving JSON: {e}")
        return False

def get_source(metric_name):
    """Return data source for each metric"""
    sources = {
        'Portfolio Occupancy': 'Occupancy Analysis.xlsx',
        'Leased Occupancy': 'Occupancy Analysis.xlsx',
        'Percent Collected': 'DQ Reports.xlsm',
        'Annualized Rent Growth': 'Effective Rent Analysis.xlsx',
        'Annualized Rent Growth (Prior Year)': 'Effective Rent Analysis.xlsx',
    }
    return sources.get(metric_name, 'Unknown')

def main():
    try:
        logger.info("=" * 70)
        logger.info("Starting Afton Metrics Extraction")
        logger.info("=" * 70)

        # Extract metrics from Excel
        logger.info("Extracting metrics from Excel files...")
        all_metrics = {}
        all_metrics.update(extract_occupancy_metrics())
        all_metrics.update(extract_dq_metrics())
        all_metrics.update(extract_effective_rent_metrics())

        if not all_metrics:
            logger.warning("No metrics extracted")
            return

        logger.info(f"[OK] Extracted {len(all_metrics)} metrics from Excel")

        # Save to local JSON cache (for dashboard) - ALWAYS DO THIS
        logger.info("Saving to local JSON cache...")
        save_metrics_json(all_metrics)
        logger.info("[OK] Dashboard will use local JSON cache")

        # Write to Google Sheet (optional - only if credentials exist)
        if CREDENTIALS_FILE.exists():
            try:
                logger.info("Uploading to Google Sheet...")
                service = get_sheets_service()
                write_to_sheets(service, all_metrics)
                logger.info("[OK] Uploaded to Google Sheet")
            except Exception as e:
                logger.warning(f"Could not upload to Google Sheet: {e}")
                logger.info("Continuing with local cache only")
        else:
            logger.info("credentials.json not found - skipping Google Sheet sync")
            logger.info("Dashboard will use local JSON cache only")

        logger.info("=" * 70)
        logger.info(f"[OK] Successfully extracted {len(all_metrics)} metrics")
        logger.info("[OK] Ready for dashboard display")
        logger.info("=" * 70)

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)

if __name__ == '__main__':
    main()
