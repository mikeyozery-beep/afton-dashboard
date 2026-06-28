#!/usr/bin/env python3
"""
Debug script to check what's in row 345 and 386 of property tabs
"""

from pathlib import Path
from openpyxl import load_workbook

GREENBOOKS_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Greenbooks")
file_path = GREENBOOKS_PATH / "Effective Rent Analysis.xlsx"

print("=" * 80)
print("DEBUGGING ROW 345 AND 386 IN RGCP TAB")
print("=" * 80)

wb = load_workbook(file_path, data_only=True)

if 'RGCP' in wb.sheetnames:
    ws = wb['RGCP']

    print("\n1. CHECKING ROW 345 (Date headers) - Searching all columns:")
    print("-" * 80)
    dates_found = []
    for col_num in range(1, 501):
        cell = ws.cell(row=345, column=col_num)
        if cell.value:
            col_letter = cell.column_letter
            dates_found.append((col_letter, col_num, cell.value))

    print(f"Found {len(dates_found)} non-empty cells in row 345")
    print(f"First 10 and last 10:")
    for col_letter, col_num, value in dates_found[:10]:
        print(f"  {col_letter}345 (col {col_num}): {value}")
    if len(dates_found) > 20:
        print("  ...")
    for col_letter, col_num, value in dates_found[-10:]:
        print(f"  {col_letter}345 (col {col_num}): {value}")

    print("\n2. CHECKING ROW 386 (Effective rent values) - Sampling columns:")
    print("-" * 80)
    values_found = []
    for col_num in range(1, 501):
        cell = ws.cell(row=386, column=col_num)
        if cell.value and isinstance(cell.value, (int, float)):
            col_letter = cell.column_letter
            values_found.append((col_letter, col_num, cell.value))

    print(f"Found {len(values_found)} numeric values in row 386")
    print(f"First 10 and last 10:")
    for col_letter, col_num, value in values_found[:10]:
        print(f"  {col_letter}386 (col {col_num}): {value:.2f}")
    if len(values_found) > 20:
        print("  ...")
    for col_letter, col_num, value in values_found[-10:]:
        print(f"  {col_letter}386 (col {col_num}): {value:.2f}")

    print("\n3. CHECKING R48 AND O48 (Month references):")
    print("-" * 80)
    print(f"  R48 (Prior year): {ws['R48'].value}")
    print(f"  O48 (Current):    {ws['O48'].value}")

else:
    print("RGCP sheet not found!")

print("\n" + "=" * 80)
