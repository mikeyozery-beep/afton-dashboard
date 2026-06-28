"""Extract budget data and calculate weighted averages"""
import openpyxl
from datetime import datetime
from pathlib import Path
import os

class BudgetExtractor:
    def __init__(self, config):
        self.config = config
        self.results = {}
        self.portfolio_index_path = None
        self.budgets_folder_path = None

    def set_paths(self, portfolio_index_path, budgets_folder_path):
        """Set paths for portfolio index and budgets folder"""
        self.portfolio_index_path = portfolio_index_path
        self.budgets_folder_path = budgets_folder_path

    def get_month_column(self, month_num):
        """Convert month number (1-12) to Excel column letter (C-M)"""
        # January=C(3), February=D(4), ..., June=H(8), ..., December=N(14)
        col_num = month_num + 2
        return chr(64 + col_num)  # Convert to column letter

    def read_portfolio_index(self):
        """Read portfolio index and return dict of {property_code: unit_count}"""
        try:
            wb = openpyxl.load_workbook(self.portfolio_index_path, read_only=True, data_only=True)
            ws = wb.active

            properties = {}
            row = 2
            while True:
                code = ws.cell(row, 2).value  # Column B
                units = ws.cell(row, 3).value  # Column C

                if not code:
                    break

                try:
                    units = int(units) if units else 0
                    properties[str(code).strip()] = units
                except:
                    pass

                row += 1

            return properties
        except Exception as e:
            print(f"Error reading portfolio index: {e}")
            return {}

    def read_property_budget(self, property_code, month_num):
        """Read occupancy budget for a property in a given month"""
        try:
            budget_file = os.path.join(
                self.budgets_folder_path,
                f"{property_code}.xlsx"
            )

            if not os.path.exists(budget_file):
                return None

            wb = openpyxl.load_workbook(budget_file, read_only=True, data_only=True)

            # Try "Assumptions" sheet
            if "Assumptions" in wb.sheetnames:
                ws = wb["Assumptions"]
            else:
                ws = wb.active

            # Row 5, column based on month
            col_letter = self.get_month_column(month_num)
            cell_ref = f"{col_letter}5"
            value = ws[cell_ref].value

            if value is None:
                return None

            try:
                return float(value)
            except:
                return None

        except Exception as e:
            print(f"Error reading budget for {property_code}: {e}")
            return None

    def calculate_occupancy_budget_weighted_avg(self, month_num=None):
        """Calculate weighted average occupancy budget for a given month"""
        try:
            if month_num is None:
                month_num = datetime.now().month

            # Get properties and unit counts
            properties = self.read_portfolio_index()
            if not properties:
                return None

            total_units = 0
            weighted_sum = 0
            properties_found = 0

            # Read each property's budget
            for property_code, units in properties.items():
                budget = self.read_property_budget(property_code, month_num)

                if budget is not None:
                    # Budget is typically stored as a percentage (95.0 not 0.95)
                    weighted_sum += budget * units
                    total_units += units
                    properties_found += 1

            if total_units == 0:
                return None

            # Calculate weighted average
            weighted_avg = weighted_sum / total_units

            self.results['occupancy_budget_weighted_avg'] = {
                'value': weighted_avg / 100.0,  # Convert to decimal
                'source': f'2026 Budgets (Weighted Average of {properties_found} properties)',
                'month': month_num,
                'timestamp': datetime.now().isoformat()
            }

            return weighted_avg / 100.0

        except Exception as e:
            print(f"Error calculating weighted average: {e}")
            return None

    def get_results(self):
        return self.results
