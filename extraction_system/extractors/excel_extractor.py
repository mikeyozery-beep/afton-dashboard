"""Extract metrics from Excel files"""
import openpyxl
from datetime import datetime
from pathlib import Path
import shutil
import tempfile
import time

class ExcelExtractor:
    def __init__(self, config):
        self.config = config
        self.results = {}

    def _read_excel_file(self, path):
        """Read Excel file, copying to temp location if needed to bypass file locks"""
        try:
            # Try reading directly first
            return openpyxl.load_workbook(path, read_only=True, data_only=True)
        except PermissionError:
            # If file is locked, copy to temp location and read from there with retry
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp_path = tmp.name

                shutil.copy2(path, tmp_path)

                # Retry with delays to handle antivirus/search locks
                for attempt in range(3):
                    try:
                        time.sleep(0.5 * (attempt + 1))  # Increase delay with each attempt
                        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                        print(f"[DEBUG] Successfully read temp copy on attempt {attempt + 1}")
                        return wb
                    except PermissionError:
                        if attempt == 2:  # Last attempt
                            raise
                        continue

            except Exception as e:
                print(f"[ERROR] Failed to read via temp copy: {e}")
                raise
            finally:
                # Clean up temp file if it exists
                if tmp_path and Path(tmp_path).exists():
                    try:
                        Path(tmp_path).unlink()
                    except:
                        pass

    def extract_occupancy_metrics(self):
        try:
            path = self.config['data_sources']['occupancy_analysis']['path']
            print(f"[DEBUG] Reading occupancy from: {path}")

            # Use temp copy method to bypass file locks
            wb = self._read_excel_file(path)
            ws = wb.active
            print(f"[DEBUG] Sheet name: {ws.title}")

            # Get cell values by reference
            portfolio_occ_current = ws['G30'].value
            portfolio_occ_prior = ws['S30'].value
            leased_occ_current = ws['H30'].value
            leased_occ_prior = ws['T30'].value

            print(f"[DEBUG] G30={portfolio_occ_current}, S30={portfolio_occ_prior}, H30={leased_occ_current}, T30={leased_occ_prior}")

            # Values are already decimals (0.96 = 96%), no conversion needed
            if portfolio_occ_current is not None:
                portfolio_occ_current = float(portfolio_occ_current)
            if portfolio_occ_prior is not None:
                portfolio_occ_prior = float(portfolio_occ_prior)
            if leased_occ_current is not None:
                leased_occ_current = float(leased_occ_current)
            if leased_occ_prior is not None:
                leased_occ_prior = float(leased_occ_prior)

            self.results['portfolio_occupancy_current'] = {
                'value': portfolio_occ_current,
                'source': 'Occupancy Analysis.xlsx, Cell G30',
                'timestamp': datetime.now().isoformat()
            }
            self.results['portfolio_occupancy_prior_week'] = {
                'value': portfolio_occ_prior,
                'source': 'Occupancy Analysis.xlsx, Cell S30',
                'timestamp': datetime.now().isoformat()
            }
            self.results['leased_occupancy_current'] = {
                'value': leased_occ_current,
                'source': 'Occupancy Analysis.xlsx, Cell H30',
                'timestamp': datetime.now().isoformat()
            }
            self.results['leased_occupancy_prior_week'] = {
                'value': leased_occ_prior,
                'source': 'Occupancy Analysis.xlsx, Cell T30',
                'timestamp': datetime.now().isoformat()
            }
            print(f"[DEBUG] Occupancy extraction complete")
            return True
        except Exception as e:
            print(f"[ERROR] Error extracting occupancy: {e}")
            import traceback
            traceback.print_exc()
            return False

    def extract_percent_collected(self):
        try:
            path = self.config['data_sources']['dq_reports']['path']
            wb = self._read_excel_file(path)
            ws = wb.active
            percent_collected = ws['H29'].value

            if percent_collected:
                percent_collected = float(percent_collected)

            self.results['percent_collected'] = {
                'value': percent_collected,
                'source': 'DQ Reports & Occupancy Summary.xlsx, Cell H29',
                'timestamp': datetime.now().isoformat()
            }
            return True
        except Exception as e:
            print(f"Error extracting collected: {e}")
            return False

    def extract_rent_growth(self):
        try:
            path = self.config['data_sources']['effective_rent_analysis']['path']
            wb = self._read_excel_file(path)
            ws = wb.active
            rent_growth = ws['O109'].value

            if rent_growth:
                rent_growth = float(rent_growth)

            self.results['annualized_rent_growth'] = {
                'value': rent_growth,
                'source': 'Effective Rent Analysis.xlsx, Cell O109',
                'timestamp': datetime.now().isoformat()
            }
            return True
        except Exception as e:
            print(f"Error extracting rent growth: {e}")
            return False

    def get_results(self):
        return self.results
