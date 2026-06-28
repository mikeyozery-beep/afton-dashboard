#!/usr/bin/env python3
"""
Download YARDI Scheduler Reports from Outlook
Extracts attachments and renames them based on workbook title + date
"""

import os
import re
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

try:
    import win32com.client
except ImportError:
    print("ERROR: pywin32 not installed")
    print("Run: pip install pywin32 --break-system-packages")
    exit(1)

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"yardi_download_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
OUTPUT_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")
OUTLOOK_INBOX = "mikereports@aftonprop.com"
SENDER_EMAIL = "cdr@yardi.com"
TEMP_DIR = SCRIPT_DIR / "temp_yardi_downloads"

# Create directories
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TEMP_DIR.mkdir(parents=True, exist_ok=True)

def get_outlook_inbox():
    """Connect to Outlook and get the specified inbox"""
    logger.info("Connecting to Outlook...")
    try:
        # Try to get existing Outlook instance
        try:
            outlook = win32com.client.GetObject(None, "Outlook.Application")
        except:
            # If that fails, try creating a new instance
            outlook = win32com.client.Dispatch("Outlook.Application")

        mapi = outlook.GetNamespace("MAPI")

        # Get the inbox folder for the specified email
        inbox = None
        for folder in mapi.Folders:
            if OUTLOOK_INBOX.lower() in folder.Name.lower():
                inbox = folder.Folders("Inbox")
                logger.info(f"[OK] Connected to inbox: {folder.Name}")
                return inbox

        # Fallback: try default inbox
        if not inbox:
            logger.warning("Could not find specific inbox, using default")
            inbox = mapi.GetDefaultFolder(6)  # 6 = olFolderInbox
            return inbox

    except Exception as e:
        logger.error(f"Error connecting to Outlook: {e}")
        raise

def get_excel_title(filepath):
    """Extract the title/name from Excel workbook"""
    try:
        wb = load_workbook(filepath, data_only=True)

        # Try to get from workbook properties
        if wb.properties.title:
            title = wb.properties.title.strip()
            logger.info(f"  Title from properties: {title}")
            return title

        # Fallback: get from first sheet name
        if wb.sheetnames:
            title = wb.sheetnames[0]
            logger.info(f"  Title from sheet name: {title}")
            return title

        # Last resort: use filename without extension
        title = Path(filepath).stem
        logger.info(f"  Using filename: {title}")
        return title

    except Exception as e:
        logger.error(f"  Error reading Excel file: {e}")
        return None

def clean_filename(name):
    """Remove invalid characters from filename"""
    # Remove invalid filename characters
    invalid_chars = r'[<>:"/\\|?*]'
    cleaned = re.sub(invalid_chars, '', name)
    # Remove extra spaces
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned

def download_and_process_attachments():
    """Main function: download attachments from Outlook"""
    logger.info("=" * 70)
    logger.info("YARDI SCHEDULER REPORT DOWNLOADER")
    logger.info("=" * 70)

    try:
        inbox = get_outlook_inbox()

        logger.info(f"\nSearching for emails from {SENDER_EMAIL}...")
        downloaded_count = 0
        error_count = 0

        # Get all emails in inbox
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)  # Sort by received time, descending

        # Process ALL emails from YARDI
        logger.info("Processing ALL emails from YARDI...")

        for i, item in enumerate(items):

            try:
                # Check if email is from YARDI
                if SENDER_EMAIL.lower() not in item.SenderEmailAddress.lower():
                    continue

                # Check if has attachments
                if item.Attachments.Count == 0:
                    continue

                email_date = item.ReceivedTime
                date_str = email_date.strftime("%m.%d.%Y")  # Format: 6.25.2026

                logger.info(f"\n[Email {i+1}] Received: {email_date.strftime('%Y-%m-%d %H:%M:%S')}")
                logger.info(f"  From: {item.SenderEmailAddress}")
                logger.info(f"  Subject: {item.Subject}")
                logger.info(f"  Attachments: {item.Attachments.Count}")

                # Process each attachment
                for attachment in item.Attachments:
                    try:
                        filename = attachment.Filename
                        logger.info(f"\n  Processing: {filename}")

                        # Skip if not Excel file
                        if not filename.lower().endswith(('.xlsx', '.xls')):
                            logger.info(f"    Skipping (not Excel): {filename}")
                            continue

                        # Save to temp location
                        temp_path = TEMP_DIR / filename
                        attachment.SaveAsFile(str(temp_path))
                        logger.info(f"    Downloaded to temp: {temp_path}")

                        # Read Excel title
                        excel_title = get_excel_title(str(temp_path))

                        if not excel_title:
                            logger.warning(f"    Could not determine title, skipping")
                            temp_path.unlink()
                            error_count += 1
                            continue

                        # Clean the title
                        clean_title = clean_filename(excel_title)

                        # Create final filename: title + date
                        final_filename = f"{clean_title} {date_str}.xlsx"
                        final_path = OUTPUT_DIR / final_filename

                        # If file exists, append number
                        if final_path.exists():
                            base = final_filename.replace('.xlsx', '')
                            counter = 1
                            while (OUTPUT_DIR / f"{base}_{counter}.xlsx").exists():
                                counter += 1
                            final_filename = f"{base}_{counter}.xlsx"
                            final_path = OUTPUT_DIR / final_filename

                        # Move from temp to final location
                        temp_path.rename(final_path)
                        logger.info(f"    ✓ Saved: {final_filename}")
                        downloaded_count += 1

                    except Exception as e:
                        logger.error(f"    Error processing attachment: {e}")
                        error_count += 1
                        continue

            except Exception as e:
                logger.error(f"  Error processing email: {e}")
                continue

        logger.info("\n" + "=" * 70)
        logger.info(f"DOWNLOAD COMPLETE")
        logger.info("=" * 70)
        logger.info(f"Downloaded: {downloaded_count} reports")
        logger.info(f"Errors: {error_count}")
        logger.info(f"Saved to: {OUTPUT_DIR}")
        logger.info("=" * 70)

        return downloaded_count, error_count

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        return 0, 1

def cleanup_temp():
    """Clean up temporary files"""
    try:
        for file in TEMP_DIR.glob("*"):
            file.unlink()
        TEMP_DIR.rmdir()
        logger.info("Cleaned up temporary files")
    except Exception as e:
        logger.warning(f"Could not clean temp directory: {e}")

if __name__ == '__main__':
    try:
        downloaded, errors = download_and_process_attachments()
        cleanup_temp()

        if errors > 0:
            exit(1)
    except KeyboardInterrupt:
        logger.warning("Interrupted by user")
        cleanup_temp()
        exit(1)
