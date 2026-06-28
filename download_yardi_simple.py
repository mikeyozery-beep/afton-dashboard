#!/usr/bin/env python3
"""
Simple YARDI Report Downloader
Direct access to Outlook inbox without complex folder navigation
"""

import os
import re
import logging
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

try:
    import win32com.client
except ImportError:
    print("ERROR: pywin32 not installed")
    exit(1)

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"yardi_download_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
OUTPUT_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")
SENDER_EMAIL = "cdr@yardi.com"
TEMP_DIR = SCRIPT_DIR / "temp_yardi_downloads"
LAST_RUN_FILE = SCRIPT_DIR / ".yardi_last_run.json"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TEMP_DIR.mkdir(parents=True, exist_ok=True)

def get_last_run_time():
    """Get the timestamp of the last successful run"""
    if LAST_RUN_FILE.exists():
        try:
            with open(LAST_RUN_FILE, 'r') as f:
                data = json.load(f)
                return datetime.fromisoformat(data['last_run'])
        except:
            return None
    return None

def save_last_run_time():
    """Save the current time as the last successful run"""
    try:
        data = {
            'last_run': datetime.now().isoformat(),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        with open(LAST_RUN_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        logger.info(f"[OK] Saved last run timestamp")
    except Exception as e:
        logger.warning(f"Could not save last run time: {e}")

def get_excel_title(filepath):
    """Extract the title/name from Excel workbook"""
    try:
        wb = load_workbook(filepath, data_only=True)
        if wb.properties.title:
            return wb.properties.title.strip()
        if wb.sheetnames:
            return wb.sheetnames[0]
        return Path(filepath).stem
    except:
        return None

def clean_filename(name):
    """Remove invalid characters from filename"""
    invalid_chars = r'[<>:"/\\|?*]'
    cleaned = re.sub(invalid_chars, '', name)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned

def download_reports(download_all=False):
    """Download YARDI reports from Outlook"""
    logger.info("=" * 70)
    logger.info("YARDI REPORT DOWNLOADER - SIMPLE")
    logger.info("=" * 70)

    last_run = get_last_run_time()
    cutoff_time = None if download_all else last_run

    if download_all:
        logger.info("\n[MODE] Download ALL reports")
    elif last_run:
        logger.info(f"\n[MODE] Download NEW reports since: {last_run.strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        logger.info("\n[MODE] First run - downloading ALL reports")

    try:
        logger.info("\nConnecting to Outlook...")
        outlook = win32com.client.GetObject(None, "Outlook.Application")
        mapi = outlook.GetNamespace("MAPI")

        # Find mikereports@aftonprop.com mailbox
        inbox = None
        for store in mapi.Stores:
            try:
                root_folder = store.GetRootFolder()
                if "mikereports" in root_folder.Name.lower() or "aftonprop" in root_folder.Name.lower():
                    inbox = root_folder.Folders("Inbox")
                    logger.info(f"[OK] Connected to: {root_folder.Name} > Inbox")
                    break
            except:
                continue

        # Fallback to default inbox if specific mailbox not found
        if not inbox:
            logger.warning("Could not find mikereports mailbox, using default Inbox")
            inbox = mapi.GetDefaultFolder(6)
            logger.info("[OK] Using default Inbox")

        logger.info(f"\nSearching for emails from {SENDER_EMAIL}...")
        items = inbox.Items
        items.Sort("[ReceivedTime]", False)  # Ascending order

        downloaded = 0
        skipped = 0
        errors = 0

        logger.info("Processing emails...")

        for i, item in enumerate(items):
            try:
                # Check sender
                if SENDER_EMAIL.lower() not in item.SenderEmailAddress.lower():
                    continue

                # Check attachments
                if item.Attachments.Count == 0:
                    continue

                email_date = item.ReceivedTime

                # Skip if before cutoff
                if cutoff_time and email_date <= cutoff_time:
                    skipped += 1
                    continue

                date_str = email_date.strftime("%m.%d.%Y")

                logger.info(f"\n[{i+1}] {email_date.strftime('%Y-%m-%d %H:%M:%S')} - {item.Subject}")
                logger.info(f"    Attachments: {item.Attachments.Count}")

                # Process attachments
                for attachment in item.Attachments:
                    filename = attachment.Filename

                    if not filename.lower().endswith(('.xlsx', '.xls')):
                        continue

                    logger.info(f"    Downloading: {filename}")

                    # Save temp
                    temp_path = TEMP_DIR / filename
                    attachment.SaveAsFile(str(temp_path))

                    # Get title
                    title = get_excel_title(str(temp_path))
                    if not title:
                        logger.warning(f"    Could not read title - skipping")
                        temp_path.unlink()
                        errors += 1
                        continue

                    # Create filename
                    clean_title = clean_filename(title)
                    final_name = f"{clean_title} {date_str}.xlsx"
                    final_path = OUTPUT_DIR / final_name

                    # Handle duplicates
                    if final_path.exists():
                        base = final_name.replace('.xlsx', '')
                        counter = 1
                        while (OUTPUT_DIR / f"{base}_{counter}.xlsx").exists():
                            counter += 1
                        final_name = f"{base}_{counter}.xlsx"
                        final_path = OUTPUT_DIR / final_name

                    # Save
                    temp_path.rename(final_path)
                    logger.info(f"    ✓ Saved: {final_name}")
                    downloaded += 1

            except Exception as e:
                logger.error(f"    Error: {e}")
                errors += 1

        logger.info("\n" + "=" * 70)
        logger.info(f"Downloaded: {downloaded} | Skipped: {skipped} | Errors: {errors}")
        logger.info(f"Saved to: {OUTPUT_DIR}")
        logger.info("=" * 70)

        # Save timestamp
        if downloaded > 0 or errors == 0:
            save_last_run_time()

        return downloaded, errors

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        return 0, 1
    finally:
        # Cleanup
        try:
            for f in TEMP_DIR.glob("*"):
                f.unlink()
        except:
            pass

if __name__ == '__main__':
    import sys
    download_all = "--all" in sys.argv
    downloaded, errors = download_reports(download_all=download_all)
    exit(0 if errors == 0 else 1)
