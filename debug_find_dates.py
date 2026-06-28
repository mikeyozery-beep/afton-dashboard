#!/usr/bin/env python3
"""
Debug script to find exact columns for 2025-06-15 and 2026-06-15
"""

from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

GREENBOOKS_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Greenbooks")
file_path = GREENBOOKS_PATH / "Effective Rent Analysis.xlsx"

print("=" * 80)
print("FINDING EXACT COLUMNS FOR TARGET DATES")
print("=" * 80)

wb = load_workbook(file_path, data_only=True)
ws = wb['RGCP']

target_dates = [
    datetime(2025, 6, 15),
    datetime(2026, 6, 15),
]

print("\nSearching for target dates in row 345...")
print("-" * 80)

for target_date in target_dates:
    found = False
    for col_num in range(1, 501):
        cell_val = ws.cell(row=345, column=col_num).value
        if cell_val and isinstance(cell_val, datetime):
            if cell_val.year == target_date.year and cell_val.month == target_date.month and cell_val.day == target_date.day:
                col_letter = ws.cell(row=345, column=col_num).column_letter
                print(f"✓ Found {target_date.date()} in column {col_letter} (col {col_num})")

                # Show the effective rent value in row 386 for this column
                rent_value = ws.cell(row=386, column=col_num).value
                print(f"  Row 386 value: {rent_value}")

                found = True
                break

    if not found:
        print(f"✗ Could not find {target_date.date()} in row 345 (columns 1-500)")

print("\n" + "=" * 80)
