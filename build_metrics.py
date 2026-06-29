#!/usr/bin/env python3
"""
Afton Properties Executive Dashboard - consolidated metrics builder.

Reads every live source (Occupancy, Collections, Effective Rent, Box Score,
Monthly Financial Reviews, Budgets, Open Positions, Weekly Meeting Notes) and
writes ONE rich JSON the dashboard consumes: data/dashboard.json

Run daily; financial sections are month-stamped (currently April 2026, the
latest completed Financial Review).
"""
import json, os, shutil, tempfile, time, traceback, re
from pathlib import Path
from datetime import datetime, date, timedelta

import openpyxl
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Paths
# ----------------------------------------------------------------------------
MO   = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery")
GB   = MO / "Greenbooks"
DD   = MO / "Dashboard Data"
BOX  = DD / "BoxScore Summary"
FIN  = MO / "5. Monthly Financial Reviews"
BUDG = MO / "2026 Budgets" / "2026 Budgets Final - Broken Links"
IDX  = MO / "Portfolio Index.xlsx"

OCC_FILE  = GB / "Occupancy Analysis.xlsx"
DQ_FILE   = GB / "DQ Reports.xlsm"
RENT_FILE = GB / "Effective Rent Analysis.xlsx"
POS_DIR   = DD / "Position"               # auto-exported timestamped Position_*.xlsx land here
OPEN_POS  = DD / "Open Positions.xlsx"     # legacy fixed-name fallback
NOTES_DIR = DD / "Weekly Meeting Notes"
PROSPECTS_DIR = DD / "rs_sql_prospects_w_reasons_muvan"   # marketing funnel source

SCRIPT_DIR = Path(__file__).parent
OUT_FILE   = SCRIPT_DIR / "data" / "dashboard.json"
HISTORY    = SCRIPT_DIR / "metrics_history.json"   # monthly snapshots (local, gitignored)
MKT_LOG    = SCRIPT_DIR / "marketing_prospects_log.json"   # running prospect log (local, gitignored)

# Which Financial Review month folder to use (latest completed). Auto-detect newest "NN. Month YYYY".
FIN_PRIOR_MONTH_STAFFING = 12   # prior-month staffing vacancy count for the delta (stored, per spec)

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def load(path):
    """Open a workbook read-only, working around Windows/OneDrive file locks."""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / (f"aftonlock_{os.getpid()}_" + path.name)
        last = None
        for attempt in range(4):
            try:
                shutil.copy2(path, tmp)
                return openpyxl.load_workbook(tmp, read_only=True, data_only=True)
            except PermissionError as e:
                last = e
                time.sleep(0.5 * (attempt + 1))
        raise last

def num(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None

def code_from(s):
    """Normalize a property code: 'ap-rgsh' / 'RGSH Final' -> 'RGSH'."""
    if not s:
        return None
    s = str(s).strip().upper()
    if s.startswith("AP-"):
        s = s[3:]
    return s.split()[0].split("/")[0].strip()

# ----------------------------------------------------------------------------
# Section extractors  (each returns plain dict; failures are isolated)
# ----------------------------------------------------------------------------
def get_occupancy():
    """Occupancy Analysis -> Summary sheet. Portfolio + per-property occ/leased."""
    wb = load(OCC_FILE); ws = wb["Summary"]
    as_of = None
    f1 = ws["F1"].value
    if isinstance(f1, str) and "as of" in f1.lower():
        as_of = f1.split("as of")[-1].strip()
    per_property = {}
    names = {}
    for r in range(7, 30):  # rows 7-29 are properties; row 30 = Total/Average
        name = ws.cell(r, 6).value           # F
        code = code_from(ws.cell(r, 4).value)  # D
        if not name or str(name).startswith("Total"):
            continue
        occ_v = num(ws.cell(r, 7).value)            # G current occupancy
        wow   = num(ws.cell(r, 9).value)            # I WoW Occ Variance (= current - prior week)
        per_property[code] = {
            "name": str(name).strip(),
            "occupancy": occ_v,
            "leased":    num(ws.cell(r, 8).value),   # H
            "trend":     num(ws.cell(r, 11).value),  # K Forecasted Occupancy
            "occ_prior_week": (occ_v - wow) if (occ_v is not None and wow is not None) else None,
        }
        names[code] = str(name).strip()
    res = {
        "portfolio_occupancy_current":    num(ws["G30"].value),
        "portfolio_occupancy_prior_week": num(ws["S30"].value),
        "leased_occupancy_current":       num(ws["H30"].value),
        "leased_occupancy_prior_week":    None,  # T30 is 'Net Occ After Leases', NOT prior-week leased; no clean source

        "trend_occupancy_30d":            num(ws["K30"].value),  # K = Forecasted/Trend Occupancy (level)
        "as_of": as_of,
        "_per_property": per_property,
        "_names": names,
    }
    wb.close()
    return res

def get_collections():
    """DQ Reports.xlsm -> Summary sheet. Portfolio H29 + per-property H."""
    wb = load(DQ_FILE); ws = wb["Summary"]
    as_of = None
    e1 = ws["E1"].value
    if isinstance(e1, str) and "as of" in e1.lower():
        as_of = e1.split("as of")[-1].strip()
    per_property = {}
    for r in range(6, 29):  # rows 6-28 props; row 29 = Total/Average
        name = ws.cell(r, 5).value            # E
        code = code_from(ws.cell(r, 3).value)   # C
        if not name or str(name).startswith("Total"):
            continue
        per_property[code] = {
            "collected": num(ws.cell(r, 8).value),       # H = % rent collected
            "prior_month": num(ws.cell(r, 52).value),    # AZ = T1 Average Rent Collected %
        }
    res = {
        "percent_collected_current": num(ws["H29"].value),
        "percent_collected_prior_month": num(ws["AZ29"].value),  # 'T1 Average Rent Collected %'
        "as_of": as_of,
        "_per_property": per_property,
    }
    wb.close()
    return res

def get_rent_growth():
    """Effective Rent Analysis -> Summary. Row 109 = YoY %, O=current, N=prior."""
    wb = load(RENT_FILE); ws = wb["Summary"]
    res = {
        "rent_growth_current": num(ws["O109"].value),
        "rent_growth_prior":   num(ws["N109"].value),
    }
    wb.close()
    return res

def _mkt_date(v):
    """Parse a prospect date cell ('6/19/2026 11:58:..' / '6/19/2026' / datetime) -> date."""
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().split()[0]
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _ingest_prospects():
    """Upsert the latest prospects export into the running log (keyed by prospect_code).
    Report header row 3, data from row 7: F=prospect_code, K=first_contact, L=show, M=application.
    Dates fill in over time, so we keep earliest first-contact and latest non-null show/app."""
    log = {}
    if MKT_LOG.exists():
        try:
            log = json.loads(MKT_LOG.read_text())
        except Exception:
            log = {}
    files = sorted(PROSPECTS_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime) if PROSPECTS_DIR.exists() else []
    if not files:
        return log
    wb = load(files[-1]); ws = wb.active
    # stream rows (read-only random ws.cell() access is O(n) per call -> far too slow here)
    for row in ws.iter_rows(min_row=7, values_only=True):
        if len(row) < 13:
            continue
        pcode = row[5]                       # F prospect_code
        if not pcode:
            continue
        key = str(pcode).strip()
        fc   = _mkt_date(row[10])            # K first_contact_date
        show = _mkt_date(row[11])            # L show_date
        app  = _mkt_date(row[12])            # M application_date
        e = log.get(key, {})
        fci = fc.isoformat() if fc else None
        if fci and (not e.get("fc") or fci < e["fc"]):
            e["fc"] = fci
        elif "fc" not in e:
            e["fc"] = fci
        if show:
            e["show"] = show.isoformat()
        elif "show" not in e:
            e["show"] = None
        if app:
            e["app"] = app.isoformat()
        elif "app" not in e:
            e["app"] = None
        e["prop"] = row[2]                   # C property_code
        log[key] = e
    wb.close()
    try:
        MKT_LOG.write_text(json.dumps(log))
    except Exception:
        pass
    return log

def _funnel(entries, today, earliest):
    """Compute the 7-day funnel + prior-4-week-avg benchmark for a set of prospect entries."""
    def week_counts(w):
        hi = today - timedelta(days=7 * w)
        lo = hi - timedelta(days=7)
        t = s = a = 0
        for e in entries:
            for fld in ("fc", "show", "app"):
                ds = e.get(fld)
                if ds:
                    d = date.fromisoformat(ds)
                    if lo < d <= hi:
                        if fld == "fc": t += 1
                        elif fld == "show": s += 1
                        else: a += 1
        return {"traffic": t, "tours": s, "applications": a, "lo": lo, "hi": hi}

    wk = {w: week_counts(w) for w in range(5)}
    cur = wk[0]
    current = {
        "traffic": cur["traffic"], "tours": cur["tours"], "applications": cur["applications"],
        "tours_over_traffic": (cur["tours"] / cur["traffic"]) if cur["traffic"] else None,
        "conversions": (cur["applications"] / cur["tours"]) if cur["tours"] else None,
    }
    prior = [wk[w] for w in range(1, 5) if earliest and wk[w]["lo"] >= earliest]

    def ratio(w, kind):
        if kind == "tours_over_traffic":
            return (w["tours"] / w["traffic"]) if w["traffic"] else None
        return (w["applications"] / w["tours"]) if w["tours"] else None

    bench_pct = {}
    for k in ("traffic", "tours", "applications"):
        vals = [w[k] for w in prior]
        avg = (sum(vals) / len(vals)) if vals else None
        bench_pct[k] = ((current[k] - avg) / avg) if avg else None
    for kind in ("tours_over_traffic", "conversions"):
        rs = [ratio(w, kind) for w in prior if ratio(w, kind) is not None]
        avg = (sum(rs) / len(rs)) if rs else None
        cv = current[kind]
        bench_pct[kind] = ((cv - avg) / avg) if (avg and cv is not None) else None
    current.update({
        "period": f"{(cur['lo'] + timedelta(days=1)).strftime('%m/%d')}-{today.strftime('%m/%d/%Y')}",
        "benchmark_pct": bench_pct,
        "benchmark_weeks": len(prior),
    })
    return current

def get_marketing(now):
    """Marketing funnel from the prospects running log: unique prospects, counted BY EVENT DATE
    (Traffic=first_contact, Tours=show, Applications=application). Past 7 days vs prior-4-week avg.
    Returns portfolio funnel plus per-property funnels under _by_property."""
    log = _ingest_prospects()
    if not log:
        return {"error": "no prospect data"}
    today = now.date()
    entries = list(log.values())
    fcs = [date.fromisoformat(e["fc"]) for e in entries if e.get("fc")]
    earliest = min(fcs) if fcs else None

    current = _funnel(entries, today, earliest)

    groups = {}
    for e in entries:
        c = code_from(e.get("prop") or "")
        if c:
            groups.setdefault(c, []).append(e)
    current["_by_property"] = {c: _funnel(es, today, earliest) for c, es in groups.items()}

    current.update({
        "log_size": len(log),
        "data_since": earliest.isoformat() if earliest else None,
    })
    return current

def latest_fin_folder():
    cands = []
    for p in FIN.iterdir():
        if p.is_dir() and p.name[:2].isdigit() and "Financial Reviews" in p.name:
            cands.append(p)
    if not cands:
        return None
    # name like "04. April 2026 Financial Reviews" -> sort by leading number within newest year
    def key(p):
        return p.stat().st_mtime
    return sorted(cands, key=key)[-1]

def _find_row(ws, gl, max_row=120):
    for r in range(1, max_row + 1):
        if str(ws.cell(r, 1).value).strip() == gl:
            return r
    return None

def get_financials(units):
    """One pass over the 23 'RGxx Final.xlsx' files. Returns portfolio aggregates AND
    per-property values for NOI var, CapEx var, and annualized-MoM rent (actual & budget).
    Distributions are raw PROPERTY COUNTS (not %)."""
    folder = latest_fin_folder()
    if folder is None:
        return {"error": "no financial review folder"}
    month_label = folder.name.split(" Financial Reviews")[0].split(". ")[-1]

    noi_ytd = noi_bud = capex_act = capex_bud = 0.0
    ni_act = ni_bud = 0.0
    noi_buckets = {"above": 0, "online": 0, "below": 0}
    ni_buckets  = {"above": 0, "at": 0, "below": 0}
    per = {}                      # code -> {noi_var, capex_var, rent_actual, rent_budget}
    ra_ws = ra_wu = rb_ws = rb_wu = 0.0
    n = 0
    for fp in sorted(folder.glob("RG* Final.xlsx")):
        code = code_from(fp.name)
        try:
            wb = load(fp); fs = wb["Financial Snapshot"]
            nj, nk, nl = num(fs.cell(83, 10).value), num(fs.cell(83, 11).value), num(fs.cell(83, 12).value)
            ij, ik, il = num(fs.cell(104, 10).value), num(fs.cell(104, 11).value), num(fs.cell(104, 12).value)
            u_j, u_k = num(fs.cell(88, 10).value), num(fs.cell(88, 11).value)
            b_j, b_k = num(fs.cell(94, 10).value), num(fs.cell(94, 11).value)
            prow = _find_row(fs, "4212-0000", 40)            # Total Potential Rent (actual)
            ra_recent = num(fs.cell(prow, 3).value) if prow else None   # C current month
            ra_prior  = num(fs.cell(prow, 6).value) if prow else None   # F prior month
            rb_recent = rb_prior = None
            if "Budget" in wb.sheetnames:
                bs = wb["Budget"]
                tcol = next((c for c in range(3, bs.max_column + 1)
                             if str(bs.cell(5, c).value).strip().lower() == "total"), None)
                brow = _find_row(bs, "4212-0000", 80)
                if tcol and brow:
                    rb_recent = num(bs.cell(brow, tcol - 1).value)
                    rb_prior  = num(bs.cell(brow, tcol - 2).value)
            wb.close()
        except Exception:
            continue
        n += 1
        if nj is not None: noi_ytd += nj
        if nk is not None: noi_bud += nk
        if ij is not None: ni_act += ij
        if ik is not None: ni_bud += ik
        cx_a = (u_j or 0) + (b_j or 0)
        cx_b = (u_k or 0) + (b_k or 0)
        capex_act += cx_a; capex_bud += cx_b
        ra = ((ra_recent / ra_prior) ** 12 - 1) if (ra_recent and ra_prior and ra_prior > 0) else None
        rb = ((rb_recent / rb_prior) ** 12 - 1) if (rb_recent and rb_prior and rb_prior > 0) else None
        per[code] = {
            "noi_var": nl,
            "noi_actual": nj, "noi_budget": nk,
            "ni_actual": ij, "ni_budget": ik, "ni_var": il,
            "capex_var": (cx_a / cx_b - 1) if cx_b else None,
            "rent_actual": ra, "rent_budget": rb,
        }
        w = units.get(code, 0)
        if ra is not None and w: ra_ws += ra * w; ra_wu += w
        if rb is not None and w: rb_ws += rb * w; rb_wu += w
        if nl is not None:
            noi_buckets["above" if nl > 0.01 else "below" if nl < -0.01 else "online"] += 1
        if il is not None:
            ni_buckets["above" if il > 0.05 else "below" if il < -0.05 else "at"] += 1

    return {
        "month_label": month_label,
        "properties_counted": n,
        "noi_variance_to_budget_ytd": (noi_ytd / noi_bud - 1) if noi_bud else None,
        "noi_actual_ytd": noi_ytd,
        "noi_budget_ytd": noi_bud,
        "ni_actual_ytd": ni_act,
        "ni_budget_ytd": ni_bud,
        "ni_variance_to_budget_ytd": (ni_act / ni_bud - 1) if ni_bud else None,
        "capex_actual_ytd": capex_act,
        "capex_budget_ytd": capex_bud,
        "capex_vs_budget_ytd": (capex_act / capex_bud - 1) if capex_bud else None,
        "rent_actual_portfolio": (ra_ws / ra_wu) if ra_wu else None,
        "rent_budget_portfolio": (rb_ws / rb_wu) if rb_wu else None,
        "noi_distribution": dict(noi_buckets),       # raw counts
        "net_income_distribution": dict(ni_buckets), # raw counts
        "_per_property": per,
        "_noi_var_by_code": {c: v["noi_var"] for c, v in per.items()},
    }

def get_units():
    """Unit counts per property code from DQ Reports 'Index' sheet (A=code, B=#units)."""
    units = {}
    try:
        wb = load(DQ_FILE); ws = wb["Index"]
        for r in range(2, ws.max_row + 1):
            code = ws.cell(r, 1).value
            u = num(ws.cell(r, 2).value)
            if not code:
                continue
            if u is not None:
                units[code_from(code)] = int(u)
        wb.close()
    except Exception:
        pass
    return units

def get_budget_targets(month_num, units):
    """Weighted-average Occupancy & Collections budget targets from per-property
    budget files. Assumptions sheet: row 5=Occupancy, row 12=Collections %.
    Month columns start at C(=Jan); column = month_num + 2."""
    col = month_num + 2  # C=3 for January
    acc = {"occupancy": [0.0, 0.0], "collected": [0.0, 0.0]}  # [weighted_sum, units]
    per = {}
    for fp in BUDG.glob("RG*.xlsx"):
        code = code_from(fp.stem)
        try:
            wb = load(fp)
            ws = wb["Assumptions"] if "Assumptions" in wb.sheetnames else wb.active
            occ = num(ws.cell(5, col).value)
            coll = num(ws.cell(12, col).value)
            wb.close()
        except Exception:
            continue
        if occ is not None and occ > 1.5:
            occ = occ / 100.0
        per[code] = {"occ_budget": occ, "coll_budget": coll}
        w = units.get(code, 0)
        if w:
            if occ is not None:
                acc["occupancy"][0] += occ * w; acc["occupancy"][1] += w
            if coll is not None:
                acc["collected"][0] += coll * w; acc["collected"][1] += w

    def wavg(key):
        s, u = acc[key]
        return (s / u) if u else None
    return {
        "occupancy_budget": wavg("occupancy"),
        "collected_budget": wavg("collected"),
        "_per_property": per,
    }

def get_rent_growth_actual(units):
    """ACTUAL annualized rent growth on the SAME basis as the budget: per property,
    take ACTUAL Total Potential Rent (GL 4212-0000) from the Financial Snapshot for the
    current month (col C) vs prior month (col F), annualize the MoM change, weight by units."""
    folder = latest_fin_folder()
    if folder is None:
        return None
    wsum = wunits = 0.0
    for fp in sorted(folder.glob("RG* Final.xlsx")):
        code = code_from(fp.name)
        w = units.get(code, 0)
        if w == 0:
            continue
        try:
            wb = load(fp); ws = wb["Financial Snapshot"]
            row = None
            for r in range(1, min(ws.max_row, 120) + 1):
                if str(ws.cell(r, 1).value).strip() == "4212-0000":
                    row = r; break
            recent = num(ws.cell(row, 3).value) if row else None   # C = Current Month
            prior  = num(ws.cell(row, 6).value) if row else None   # F = Prior Month
            wb.close()
        except Exception:
            continue
        if recent and prior and prior > 0:
            wsum += ((recent / prior) ** 12 - 1) * w; wunits += w
    return (wsum / wunits) if wunits else None

def get_rent_growth_budget(units):
    """Annualized budgeted rent growth: per property, take Total Potential Rent
    (GL 4212-0000) from the Budget tab of each financial-review file for the most
    recent month vs the prior month, annualize the MoM change, then weight by units."""
    folder = latest_fin_folder()
    if folder is None:
        return None
    wsum = wunits = 0.0
    for fp in sorted(folder.glob("RG* Final.xlsx")):
        code = code_from(fp.name)
        w = units.get(code, 0)
        if w == 0:
            continue
        try:
            wb = load(fp)
            if "Budget" not in wb.sheetnames:
                wb.close(); continue
            ws = wb["Budget"]
            # locate the "Total" column in header row 5 -> recent = Total-1, prior = Total-2
            total_col = None
            for c in range(3, ws.max_column + 1):
                if str(ws.cell(5, c).value).strip().lower() == "total":
                    total_col = c; break
            # locate Total Potential Rent row (GL 4212-0000)
            pot_row = None
            for r in range(6, min(ws.max_row, 80) + 1):
                if str(ws.cell(r, 1).value).strip() == "4212-0000":
                    pot_row = r; break
            wb_recent = wb_prior = None
            if total_col and pot_row:
                wb_recent = num(ws.cell(pot_row, total_col - 1).value)
                wb_prior  = num(ws.cell(pot_row, total_col - 2).value)
            wb.close()
        except Exception:
            continue
        if wb_recent and wb_prior and wb_prior > 0:
            annual = (wb_recent / wb_prior) ** 12 - 1   # compound-annualized MoM change
            wsum += annual * w; wunits += w
    return (wsum / wunits) if wunits else None

def get_staffing():
    """Live count of Status==Open in the SharePoint 'Open Positions' list.
    Tries Microsoft Graph (delegated) first; falls back to the latest local export."""
    # 1) live via Graph (requires one-time `python graph_staffing.py login`)
    try:
        import graph_staffing
        g = graph_staffing.open_positions_count()
        if g and g.get("value") is not None:
            return {"value": g["value"], "prior_month": FIN_PRIOR_MONTH_STAFFING,
                    "source": g.get("source", "SharePoint (Graph)")}
    except Exception as e:
        log_graph = f"graph staffing unavailable ({e}); using local export"
        print(log_graph)

    # 2) fallback: latest Position_*.xlsx export
    src = None
    if POS_DIR.exists():
        files = sorted(POS_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
        if files:
            src = files[-1]
    if src is None and OPEN_POS.exists():
        src = OPEN_POS
    if src is None:
        return {"value": None, "status": "FILE_NOT_FOUND"}
    wb = load(src); ws = wb.active
    # find Status column from header row
    status_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).strip().lower() == "status":
            status_col = c; break
    open_count = 0
    if status_col:
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, status_col).value
            if isinstance(v, str) and v.strip().lower() == "open":
                open_count += 1
    wb.close()
    return {"value": open_count, "prior_month": FIN_PRIOR_MONTH_STAFFING,
            "source": f"{src.name} (SharePoint export)"}

def get_staffing_by_property():
    """Open-position count per property code from the latest local Position export.
    Title column (B) holds the property code (occasionally 'RGAL/RGST' -> split)."""
    counts = {}
    src = None
    if POS_DIR.exists():
        files = sorted(POS_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
        if files:
            src = files[-1]
    if src is None and OPEN_POS.exists():
        src = OPEN_POS
    if src is None:
        return counts
    wb = load(src); ws = wb.active
    hdr = {str(ws.cell(1, c).value).strip().lower(): c for c in range(1, ws.max_column + 1)}
    sc = hdr.get("status"); tc = hdr.get("title")
    if sc and tc:
        for r in range(2, ws.max_row + 1):
            st = ws.cell(r, sc).value
            if isinstance(st, str) and st.strip().lower() == "open":
                title = str(ws.cell(r, tc).value or "")
                for part in title.split("/"):
                    code = code_from(part)
                    if code:
                        counts[code] = counts.get(code, 0) + 1
    wb.close()
    return counts

def get_takeaways():
    """Latest .docx in Weekly Meeting Notes -> list-paragraph bullets."""
    try:
        import docx
    except ImportError:
        return {"items": [], "error": "python-docx not installed"}
    docs = sorted(NOTES_DIR.glob("*.docx"), key=lambda p: p.stat().st_mtime)
    docs = [d for d in docs if not d.name.startswith("~$")]
    if not docs:
        return {"items": [], "error": "no notes found"}
    d = docx.Document(str(docs[-1]))
    bullets = []
    for p in d.paragraphs:
        t = p.text.strip()
        if not t:
            continue
        if "List" in (p.style.name or "") or t.startswith(("•", "-", "*")):
            bullets.append(t.lstrip("•-* ").strip())
    return {"items": bullets, "source_file": docs[-1].name}

def get_areas_of_focus():
    """Three portfolio-average focus themes from Effective Rent Analysis -> Summary:
    Tradeouts (block from row 4), Renewals (block from row 30), Effective-rent MoM (O108).
    Each is a unit-weighted portfolio average shown as current month vs prior month; the
    section as-of (e.g. 'June vs May 2026') is returned so bullets don't repeat the month."""
    wb = load(RENT_FILE); ws = wb["Summary"]

    def month_cols(hdr_row):
        return [c for c in range(4, 32) if isinstance(ws.cell(hdr_row, c).value, datetime)]

    def weighted_avg(data_start, col):
        if not col:
            return None
        wsum = wunits = 0.0
        for r in range(data_start, data_start + 30):
            code = ws.cell(r, 3).value
            if not code or not str(code).strip().upper().startswith("RG"):
                break
            val = num(ws.cell(r, col).value)
            u   = num(ws.cell(r, 1).value)
            if val is not None and u:
                wsum += val * u; wunits += u
        return (wsum / wunits) if wunits else None

    def cur_prior(data_start, hdr_row):
        cols = month_cols(hdr_row)
        last  = cols[-1] if cols else None
        prior = cols[-2] if len(cols) >= 2 else None
        return (weighted_avg(data_start, last), weighted_avg(data_start, prior),
                ws.cell(hdr_row, last).value if last else None,
                ws.cell(hdr_row, prior).value if prior else None)

    trade, trade_prior, t_asof, t_prior_asof = cur_prior(4, 3)     # Tradeouts (header row 3)
    renew, renew_prior, _, _                 = cur_prior(30, 29)   # Renewals  (header row 29)
    eff_mom       = num(ws["O108"].value)      # Effective Rents 'MoM %' latest month (June)
    eff_mom_prior = num(ws["N108"].value)      # prior month (May)
    wb.close()

    def mname(d):
        try:
            return d.strftime("%B")
        except Exception:
            return ""
    cur_m, prior_m = mname(t_asof), mname(t_prior_asof)
    try:
        yr = t_asof.year
    except Exception:
        yr = ""
    as_of = (f"{cur_m} vs {prior_m} {yr}".strip() if cur_m and prior_m
             else (f"{cur_m} {yr}".strip() if cur_m else ""))

    def cur_vs_prior(label, cur, prior, suffix=""):
        if cur is None:
            return f"{label}: no current data"
        s = f"{label} averaging {cur*100:+.1f}%{suffix}"
        if prior is not None:
            s += f" vs {prior*100:+.1f}%{suffix} in prior month"
        return s

    focus = []
    focus.append({"theme": "Tradeouts", "summary": cur_vs_prior("Tradeouts", trade, trade_prior)})
    focus.append({"theme": "Renewals",  "summary": cur_vs_prior("Renewals", renew, renew_prior)})
    annz = lambda m: ((1 + m) ** 12 - 1) if m is not None else None
    eff_cur, eff_prior = annz(eff_mom), annz(eff_mom_prior)
    if eff_cur is None:
        eff_txt = "Effective rent: no current data"
    else:
        eff_txt = f"Effective rent {eff_cur*100:+.1f}% annualized"
        if eff_prior is not None:
            eff_txt += f" vs {eff_prior*100:+.1f}% annualized in prior month"
    focus.append({"theme": "Effective Rent", "summary": eff_txt})
    return {"as_of": as_of, "items": focus}


# ----------------------------------------------------------------------------
# Analysis: bottom performers + focus areas (anomalies)
# ----------------------------------------------------------------------------
def build_analysis(occ, coll, fin):
    names = occ["_names"]
    occ_pp  = occ["_per_property"]
    coll_pp = coll["_per_property"]
    noi_pp  = fin.get("_noi_var_by_code", {})

    rows = []
    for code, name in names.items():
        o = occ_pp.get(code, {})
        rows.append({
            "code": code, "name": name,
            "occupancy": o.get("occupancy"),
            "leased": o.get("leased"),
            "collected": (coll_pp.get(code) or {}).get("collected"),
            "noi_variance": noi_pp.get(code),
        })

    # composite score (tiebreaker) across available metrics. Normalize each metric 0..1.
    metrics = ["occupancy", "leased", "collected", "noi_variance"]
    ranges = {}
    for m in metrics:
        vals = [r[m] for r in rows if r[m] is not None]
        ranges[m] = (min(vals), max(vals)) if vals else (0, 1)
    for r in rows:
        parts = []
        for m in metrics:
            lo, hi = ranges[m]
            if r[m] is None or hi == lo:
                continue
            parts.append((r[m] - lo) / (hi - lo))
        r["score"] = sum(parts) / len(parts) if parts else 1.0

    # Occupancy-led: rank primarily by occupancy (lowest = worst); composite breaks ties.
    bottom = sorted(rows, key=lambda r: (r["occupancy"] if r["occupancy"] is not None else 1.0,
                                         r["score"]))[:3]

    # Most-concerning column = the metric furthest BELOW its portfolio benchmark.
    # occ/leased/collected benchmark = portfolio current; NOI variance benchmark = 0 (budget).
    bench = {
        "occupancy":    occ.get("portfolio_occupancy_current"),
        "leased":       occ.get("leased_occupancy_current"),
        "collected":    coll.get("percent_collected_current"),
        "noi_variance": 0.0,
    }
    def worst_col(r):
        gaps = {}
        for m, b in bench.items():
            v = r.get(m)
            if v is None or b is None:
                continue
            gaps[m] = b - v          # positive => below benchmark => worse
        if not gaps:
            return None
        m = max(gaps, key=gaps.get)
        return m if gaps[m] > 0 else None

    bottom_out = [{
        "name": r["name"],
        "occupancy": r["occupancy"], "leased": r["leased"],
        "collected": r["collected"], "noi_variance": r["noi_variance"],
        "rent_growth": None,  # not available per-property in a single cell
        "concern": worst_col(r),
    } for r in bottom]

    # Focus areas = occupancy anomalies (lowest occupancy properties)
    occ_sorted = sorted([r for r in rows if r["occupancy"] is not None],
                        key=lambda r: r["occupancy"])[:3]
    sev = ["High", "Medium", "Medium"]
    focus = [{
        "rank": i + 1,
        "title": "Occupancy Concern",
        "summary": f"{r['name']} occupancy at {r['occupancy']*100:.1f}%",
        "property": r["name"],
        "severity": sev[i] if i < len(sev) else "Medium",
    } for i, r in enumerate(occ_sorted)]

    return bottom_out, focus

def build_views(occ, coll, bud, fin, staff, staff_pp, mkt, prior):
    """Assemble the toggleable rows 1-3 for Portfolio + each property."""
    names   = occ.get("_names", {})
    occ_pp  = occ.get("_per_property", {})
    coll_pp = coll.get("_per_property", {})
    bud_pp  = bud.get("_per_property", {})
    fin_pp  = fin.get("_per_property", {})
    mkt_pp  = mkt.get("_by_property", {}) if isinstance(mkt, dict) else {}

    def mkt_view(f):
        if not f:
            return {}
        return {k: f.get(k) for k in ("traffic", "tours", "applications", "tours_over_traffic",
                                      "conversions", "period", "benchmark_pct", "benchmark_weeks")}

    # single-property bucket (matches the portfolio distribution thresholds in get_financials)
    def noi_bucket(v):
        if v is None:
            return {"above": 0, "online": 0, "below": 0}
        return {"above": int(v > 0.01), "online": int(-0.01 <= v <= 0.01), "below": int(v < -0.01)}
    def ni_bucket(v):
        if v is None:
            return {"above": 0, "at": 0, "below": 0}
        return {"above": int(v > 0.05), "at": int(-0.05 <= v <= 0.05), "below": int(v < -0.05)}

    views = {
        "PORTFOLIO": {
            "label": f"Portfolio ({fin.get('properties_counted') or len(names)} properties)",
            "kpis": {
                "portfolio_occupancy": {"value": occ.get("portfolio_occupancy_current"),
                                        "prior_week": occ.get("portfolio_occupancy_prior_week"),
                                        "budget": bud.get("occupancy_budget")},
                "leased_occupancy": {"value": occ.get("leased_occupancy_current"), "budget": None},
                "percent_collected": {"value": coll.get("percent_collected_current"),
                                      "prior_month": coll.get("percent_collected_prior_month"),
                                      "budget": bud.get("collected_budget")},
                "rent_growth": {"value": fin.get("rent_actual_portfolio"),
                                "budget": fin.get("rent_budget_portfolio"),
                                "as_of": fin.get("month_label")},
                "trend_occupancy_30d": {"value": occ.get("trend_occupancy_30d"),
                                        "budget": bud.get("occupancy_budget")},
                "staffing_vacancies": {"value": staff.get("value"),
                                       "prior_month": prior.get("staffing_open")},
                "noi_variance": {"value": fin.get("noi_variance_to_budget_ytd"),
                                 "actual": fin.get("noi_actual_ytd"),
                                 "budget": fin.get("noi_budget_ytd"),
                                 "as_of": fin.get("month_label")},
                "capex_vs_budget": {"value": fin.get("capex_vs_budget_ytd"),
                                    "actual": fin.get("capex_actual_ytd"),
                                    "budget": fin.get("capex_budget_ytd"),
                                    "as_of": fin.get("month_label")},
            },
            "charts": {
                "noi": {"actual": fin.get("noi_actual_ytd"), "budget": fin.get("noi_budget_ytd"),
                        "variance": fin.get("noi_variance_to_budget_ytd"),
                        "dist": fin.get("noi_distribution")},
                "net_income": {"actual": fin.get("ni_actual_ytd"), "budget": fin.get("ni_budget_ytd"),
                               "variance": fin.get("ni_variance_to_budget_ytd"),
                               "dist": fin.get("net_income_distribution")},
                "as_of": fin.get("month_label"),
            },
            "marketing": mkt_view(mkt),
        }
    }
    for code, nm in names.items():
        o = occ_pp.get(code, {}); c = coll_pp.get(code, {})
        b = bud_pp.get(code, {});  f = fin_pp.get(code, {})
        views[code] = {
            "label": nm,
            "kpis": {
                "portfolio_occupancy": {"value": o.get("occupancy"),
                                        "prior_week": o.get("occ_prior_week"),
                                        "budget": b.get("occ_budget")},
                "leased_occupancy": {"value": o.get("leased"), "budget": None},
                "percent_collected": {"value": c.get("collected"),
                                      "prior_month": c.get("prior_month"),
                                      "budget": b.get("coll_budget")},
                "rent_growth": {"value": f.get("rent_actual"), "budget": f.get("rent_budget"),
                                "as_of": fin.get("month_label")},
                "trend_occupancy_30d": {"value": o.get("trend"), "budget": b.get("occ_budget")},
                "staffing_vacancies": {"value": staff_pp.get(code, 0), "prior_month": None},
                "noi_variance": {"value": f.get("noi_var"),
                                 "actual": f.get("noi_actual"), "budget": f.get("noi_budget"),
                                 "as_of": fin.get("month_label")},
                "capex_vs_budget": {"value": f.get("capex_var"), "actual": None,
                                    "budget": None, "as_of": fin.get("month_label")},
            },
            "charts": {
                "noi": {"actual": f.get("noi_actual"), "budget": f.get("noi_budget"),
                        "variance": f.get("noi_var"), "dist": noi_bucket(f.get("noi_var"))},
                "net_income": {"actual": f.get("ni_actual"), "budget": f.get("ni_budget"),
                               "variance": ((f["ni_actual"] / f["ni_budget"] - 1)
                                            if f.get("ni_actual") is not None and f.get("ni_budget") else None),
                               "dist": ni_bucket((f["ni_actual"] / f["ni_budget"] - 1)
                                                 if f.get("ni_actual") is not None and f.get("ni_budget") else None)},
                "as_of": fin.get("month_label"),
            },
            "marketing": mkt_view(mkt_pp.get(code)),
        }
    return views

# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def update_history(now, current):
    """Persist this month's metric snapshot; return the most-recent PRIOR month's snapshot.
    Builds month-over-month comparisons going forward (no reliable historical source exists)."""
    key = f"{now.year:04d}-{now.month:02d}"
    hist = {}
    if HISTORY.exists():
        try:
            hist = json.loads(HISTORY.read_text())
        except Exception:
            hist = {}
    prior_keys = sorted(k for k in hist if k < key)
    prior = hist.get(prior_keys[-1], {}) if prior_keys else {}
    hist.setdefault(key, {})
    hist[key].update({k: v for k, v in current.items() if v is not None})
    try:
        HISTORY.write_text(json.dumps(hist, indent=2))
    except Exception:
        pass
    return prior


def safe(fn, label, errors):
    try:
        return fn()
    except Exception as e:
        errors.append(f"{label}: {e}")
        traceback.print_exc()
        return {}

def main():
    errors = []
    now = datetime.now()

    occ   = safe(get_occupancy,    "occupancy",   errors)
    coll  = safe(get_collections,  "collections", errors)
    rent  = safe(get_rent_growth,  "rent_growth", errors)
    box   = safe(lambda: get_marketing(now), "marketing", errors)
    units = safe(get_units,        "units",       errors) or {}
    fin   = safe(lambda: get_financials(units), "financials", errors)
    bud   = safe(lambda: get_budget_targets(now.month, units), "budgets", errors)
    staff = safe(get_staffing,     "staffing",    errors)
    staff_pp = safe(get_staffing_by_property, "staffing_by_property", errors) or {}
    take  = safe(get_takeaways,    "takeaways",   errors)

    # Month-over-month comparisons (auto-snapshot; appears once a new month rolls over)
    prior = update_history(now, {"staffing_open": staff.get("value")})

    bottom = []
    if occ and coll:
        try:
            bottom, _ = build_analysis(occ, coll, fin or {})
        except Exception as e:
            errors.append(f"analysis: {e}"); traceback.print_exc()
    focus_raw = safe(get_areas_of_focus, "focus_areas", errors) or {}
    if isinstance(focus_raw, dict):
        focus, focus_as_of = focus_raw.get("items", []), focus_raw.get("as_of")
    else:
        focus, focus_as_of = focus_raw, None

    views = build_views(occ, coll, bud, fin, staff, staff_pp, box, prior)
    portfolio = views["PORTFOLIO"]
    portfolio["kpis"]["rent_growth"]["yoy_effective"] = rent.get("rent_growth_current")

    out = {
        "generated_at": now.isoformat(),
        "as_of": {
            "occupancy":   occ.get("as_of"),
            "collections": coll.get("as_of"),
            "financials":  fin.get("month_label"),
            "boxscore":    box.get("period"),
        },
        # toggleable rows 1-3 per view; HTML defaults to PORTFOLIO
        "views": views,
        "default_view": "PORTFOLIO",
        # portfolio mirror (rows 4+ and any non-toggled consumers)
        "kpis": portfolio["kpis"],
        "marketing": portfolio["marketing"],
        "bottom_performers": bottom,
        "focus_areas": focus,
        "focus_as_of": focus_as_of,
        "noi_distribution": fin.get("noi_distribution"),          # raw property counts
        "net_income_distribution": fin.get("net_income_distribution"),
        "takeaways": take.get("items", []),
        "errors": errors,
        "status": "complete" if not errors else "complete_with_warnings",
    }

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT_FILE}")
    if errors:
        print("WARNINGS:")
        for e in errors:
            print("  -", e)
    return out

if __name__ == "__main__":
    main()
