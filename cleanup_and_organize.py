#!/usr/bin/env python3
"""
Comprehensive cleanup and organization of Dashboard Data folder
1. Extracts Scheduler_Reports worksheets
2. Handles generic-named files (Report One, etc.)
3. Organizes ALL reports into folders by report name
4. Deletes original files
"""

import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import shutil

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"cleanup_organize_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")

def sanitize_filename(name):
    """Remove/replace invalid filename characters"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()

def get_report_name_from_file(filepath):
    """
    Get the actual report name from an Excel file
    Reads the first worksheet name as the report name
    """
    try:
        wb = load_workbook(filepath)
        if wb.sheetnames:
            # Use first sheet name as report name
            return wb.sheetnames[0]
    except:
        pass
    return None

def extract_and_organize_file(filepath):
    """
    Process a single file:
    - If multi-sheet: extract each sheet as separate file
    - If single-sheet: move to folder based on sheet name
    Returns list of (report_name, output_path) tuples
    """
    results = []

    try:
        wb = load_workbook(filepath)
        logger.info(f"\n[Processing] {filepath.name}")
        logger.info(f"  Worksheets: {len(wb.sheetnames)}")

        # Extract each worksheet
        for sheet_name in wb.sheetnames:
            try:
                logger.info(f"  Extracting: '{sheet_name}'")

                # Create new workbook with just this sheet
                new_wb = load_workbook(filepath)

                # Keep only this sheet
                sheets_to_delete = [s for s in new_wb.sheetnames if s != sheet_name]
                for sheet_to_delete in sheets_to_delete:
                    del new_wb[sheet_to_delete]

                # Rename sheet to generic name
                new_ws = new_wb.active
                new_ws.title = "Report"

                # Create folder for this report type
                safe_sheet_name = sanitize_filename(sheet_name)
                report_folder = DASHBOARD_DATA_DIR / safe_sheet_name
                report_folder.mkdir(exist_ok=True)

                # Create output filename with timestamp
                timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                output_filename = f"{safe_sheet_name}_{timestamp}.xlsx"
                output_path = report_folder / output_filename

                # Save the individual report
                new_wb.save(output_path)
                logger.info(f"    ✓ Saved to {safe_sheet_name}/{output_filename}")
                results.append((safe_sheet_name, output_path))

            except Exception as e:
                logger.error(f"    ✗ Error extracting '{sheet_name}': {e}")
                continue

    except Exception as e:
        logger.error(f"  Error processing {filepath.name}: {e}")

    return results

def cleanup_and_organize():
    """Main function: organize all reports into folders"""
    logger.info("=" * 70)
    logger.info("DASHBOARD DATA CLEANUP & ORGANIZATION")
    logger.info("=" * 70)
    logger.info(f"Directory: {DASHBOARD_DATA_DIR}\n")

    if not DASHBOARD_DATA_DIR.exists():
        logger.error(f"Dashboard Data folder not found")
        return

    # Find all Excel files in root
    root_excel_files = list(DASHBOARD_DATA_DIR.glob("*.xlsx")) + list(DASHBOARD_DATA_DIR.glob("*.xls"))

    if not root_excel_files:
        logger.warning("No Excel files found in root of Dashboard Data folder")
        return

    logger.info(f"Found {len(root_excel_files)} files in root folder:\n")
    for f in sorted(root_excel_files):
        logger.info(f"  - {f.name}")

    logger.info("\n" + "=" * 70)
    logger.info("EXTRACTING AND ORGANIZING")
    logger.info("=" * 70)

    total_extracted = 0
    files_to_delete = []
    report_count = {}

    # Process each file
    for excel_file in sorted(root_excel_files):
        results = extract_and_organize_file(excel_file)
        total_extracted += len(results)

        # Track report types
        for report_name, _ in results:
            report_count[report_name] = report_count.get(report_name, 0) + 1

        # Mark original for deletion
        if results:
            files_to_delete.append(excel_file)

    # Delete original files
    logger.info("\n" + "=" * 70)
    logger.info("DELETING ORIGINAL FILES")
    logger.info("=" * 70)

    for file_to_delete in files_to_delete:
        try:
            logger.info(f"  Deleting: {file_to_delete.name}")
            file_to_delete.unlink()
        except Exception as e:
            logger.error(f"  Error deleting {file_to_delete.name}: {e}")

    # Summary
    logger.info("\n" + "=" * 70)
    logger.info("ORGANIZATION COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Files processed: {len(root_excel_files)}")
    logger.info(f"Individual reports extracted: {total_extracted}")
    logger.info(f"Original files deleted: {len(files_to_delete)}")
    logger.info(f"\nReport Folders Created:")
    for report_name in sorted(report_count.keys()):
        logger.info(f"  - {report_name}/ ({report_count[report_name]} files)")
    logger.info("\nFolder Structure:")
    logger.info("  Dashboard Data/")
    logger.info("    ├── Unit_Occupancy_6413722/")
    logger.info("    │   ├── Unit_Occupancy_6413722_2026-06-20.xlsx")
    logger.info("    │   ├── Unit_Occupancy_6413722_2026-06-21.xlsx")
    logger.info("    │   └── ...")
    logger.info("    ├── ResAnalytics_Aged_Receivables/")
    logger.info("    ├── rs_sql_evictions_am_reports/")
    logger.info("    └── [Other Reports]/")
    logger.info("\n✓ All reports now organized in folders by type")
    logger.info("=" * 70)

if __name__ == '__main__':
    cleanup_and_organize()
