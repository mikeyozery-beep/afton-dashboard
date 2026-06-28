#!/usr/bin/env python3
"""
Debug script to check Effective Rent Analysis data
"""

from pathlib import Path
from openpyxl import load_workbook

GREENBOOKS_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Greenbooks")
file_path = GREENBOOKS_PATH / "Effective Rent Analysis.xlsx"

print("=" * 80)
print("DEBUGGING EFFECTIVE RENT ANALYSIS")
print("=" * 80)

wb = load_workbook(file_path, data_only=True)
ws_summary = wb['Summary']

print("\n1. CHECKING SUMMARY SHEET ROW 3 (Unit counts)")
print("-" * 80)
print("First 25 columns of row 3:")
for col_num in range(1, 26):
    cell = ws_summary.cell(row=3, column=col_num)
    col_letter = ws_summary.cell(row=3, column=col_num).column_letter
    print(f"  {col_letter}3: {cell.value}")

print("\n2. CHECKING ROW 386 IN FIRST PROPERTY TAB (RGCP)")
print("-" * 80)
if 'RGCP' in wb.sheetnames:
    ws_prop = wb['RGCP']
    print(f"Sheet RGCP exists")
    print(f"Row 386, Column A-E:")
    for col_num in range(1, 6):
        cell = ws_prop.cell(row=386, column=col_num)
        col_letter = ws_prop.cell(row=386, column=col_num).column_letter
        print(f"  {col_letter}386: {cell.value}")
else:
    print("RGCP sheet not found!")

print("\n3. ALL PROPERTY SHEET NAMES")
print("-" * 80)
print(f"Total sheets: {len(wb.sheetnames)}")
property_tabs = ['RGCP', 'RGSR', 'RGP', 'RGSH', 'RGLA', 'RGD', 'RGOX', 'RGPT',
                'RGSV', 'RGSC', 'RGTCC', 'RGCCV', 'RGCK', 'RGWL', 'RGOV', 'RGSM',
                'RGLT', 'RGST', 'RGAL', 'RGCW', 'RGAZ', 'RGHC', 'RGLSV']

found_count = 0
for prop in property_tabs:
    if prop in wb.sheetnames:
        found_count += 1
        print(f"  ✓ {prop}")
    else:
        print(f"  ✗ {prop} NOT FOUND")

print(f"\nFound {found_count} out of {len(property_tabs)} expected property tabs")

print("\n4. CURRENT VALUE IN O109 (Latest Annualized Rent Growth)")
print("-" * 80)
current = ws_summary['O109'].value
print(f"O109: {current}")

print("\n" + "=" * 80)
