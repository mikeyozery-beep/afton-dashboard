#!/usr/bin/env python3
"""
Enhanced Afton Dashboard Metrics Extraction
Extracts all metrics from multiple Excel sources and caches to local JSON
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime
from glob import glob

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
DATA_DIR = SCRIPT_DIR / "data"
LOG_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"metrics_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

METRICS_FILE = DATA_DIR / 'metrics.json'

# File paths
GREENBOOKS_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Greenbooks")
FINANCIAL_PATH = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\5. Monthly Financial Reviews")

# Property tabs in Effective Rent Analysis
PROPERTY_TABS = [
    'RGCP', 'RGSR', 'RGP', 'RGSH', 'RGLA', 'RGD', 'RGOX', 'RGPT',
    'RGSV', 'RGSC', 'RGTCC', 'RGCCV', 'RGCK', 'RGWL', 'RGOV', 'RGSM',
    'RGLT', 'RGST', 'RGAL', 'RGCW', 'RGAZ', 'RGHC', 'RGLSV'
]

def extract_occupancy_metrics():
    """Extract Portfolio Occupancy and Leased Occupancy from Occupancy Analysis"""
    try:
        file_path = GREENBOOKS_PATH / "Occupancy Analysis.xlsx"
        logger.info(f"Reading Occupancy Analysis from {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws = wb['Summary']

        portfolio_occ = ws['G30'].value
        leased_occ = ws['H30'].value

        logger.info(f"[OK] Portfolio Occupancy: {portfolio_occ}")
        logger.info(f"[OK] Leased Occupancy: {leased_occ}")

        return {
            'Portfolio Occupancy': format_percentage(portfolio_occ),
            'Leased Occupancy': format_percentage(leased_occ)
        }
    except Exception as e:
        logger.error(f"Error extracting occupancy metrics: {e}")
        return {}

def extract_dq_metrics():
    """Extract Percent Collected from DQ Reports"""
    try:
        file_path = GREENBOOKS_PATH / "DQ Reports.xlsm"
        logger.info(f"Reading DQ Reports from {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws = wb['Summary']

        percent_collected = ws['H29'].value

        logger.info(f"[OK] Percent Collected: {percent_collected}")

        return {
            'Percent Collected': format_percentage(percent_collected)
        }
    except Exception as e:
        logger.error(f"Error extracting DQ metrics: {e}")
        return {}

def extract_effective_rent_metrics():
    """Extract Annualized Rent Growth (current and prior year)"""
    try:
        file_path = GREENBOOKS_PATH / "Effective Rent Analysis.xlsx"
        logger.info(f"Reading Effective Rent Analysis from {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws_summary = wb['Summary']

        # Current month (Latest)
        current_growth = ws_summary['O109'].value
        logger.info(f"[OK] Annualized Rent Growth (Current): {current_growth}")

        # Prior Year (SUMPRODUCT calculation)
        prior_year_growth = calculate_prior_year_growth(wb, ws_summary)
        logger.info(f"[OK] Annualized Rent Growth (Prior Year): {prior_year_growth}")

        return {
            'Annualized Rent Growth': format_percentage(current_growth),
            'Annualized Rent Growth (Prior Year)': format_percentage(prior_year_growth)
        }
    except Exception as e:
        logger.error(f"Error extracting effective rent metrics: {e}")
        return {}

def calculate_prior_year_growth(wb, ws_summary):
    """Calculate weighted average effective rent growth using SUMPRODUCT logic"""
    try:
        from datetime import datetime, timedelta

        logger.info("Calculating prior year annualized rent growth (SUMPRODUCT)...")

        weighted_sum = 0
        total_units = 0

        for prop_code in PROPERTY_TABS:
            if prop_code not in wb.sheetnames:
                logger.warning(f"  {prop_code}: Sheet not found")
                continue

            ws_prop = wb[prop_code]

            # Read prior year month from R48
            prior_month = ws_prop['R48'].value
            if not prior_month:
                logger.warning(f"  {prop_code}: R48 is empty")
                continue

            # Find column with prior year date in row 345 (search up to column BZ and beyond)
            prior_year_col = None
            for col_num in range(1, 500):  # Search up to column 500 (well beyond BZ)
                cell_val = ws_prop.cell(row=345, column=col_num).value
                if cell_val and isinstance(cell_val, datetime):
                    if cell_val.year == prior_month.year and cell_val.month == prior_month.month:
                        prior_year_col = col_num
                        logger.info(f"  {prop_code}: Found prior year date {cell_val} in column {col_num}")
                        break

            if not prior_year_col:
                logger.warning(f"  {prop_code}: Could not find date {prior_month} in row 345 (searched columns 1-500)")
                continue

            # Read unit count from B386
            unit_count = ws_prop['B386'].value
            if not unit_count or not isinstance(unit_count, (int, float)):
                logger.warning(f"  {prop_code}: Invalid unit count in B386: {unit_count}")
                continue

            # Read prior year effective rent from row 386 of found column
            prior_year_rent = ws_prop.cell(row=386, column=prior_year_col).value
            if prior_year_rent and isinstance(prior_year_rent, (int, float)):
                weighted_sum += unit_count * prior_year_rent
                total_units += unit_count
                logger.info(f"  {prop_code}: {unit_count} units x {prior_year_rent:.6f}")

        if total_units > 0:
            weighted_average_prior = weighted_sum / total_units
            logger.info(f"  Prior year weighted average rent: ${weighted_average_prior:.2f}")

            # Get current month weighted average from O48 column
            # We need to recalculate current month the same way
            logger.info("Calculating current month weighted average for growth rate...")
            current_weighted_sum = 0
            current_total_units = 0

            for prop_code in PROPERTY_TABS:
                if prop_code not in wb.sheetnames:
                    continue

                ws_prop = wb[prop_code]
                current_month = ws_prop['O48'].value

                if not current_month:
                    continue

                # Find column with current month date in row 345
                current_col = None
                for col_num in range(1, 500):
                    cell_val = ws_prop.cell(row=345, column=col_num).value
                    if cell_val and isinstance(cell_val, datetime):
                        if cell_val.year == current_month.year and cell_val.month == current_month.month:
                            current_col = col_num
                            break

                if not current_col:
                    continue

                unit_count = ws_prop['B386'].value
                if not unit_count or not isinstance(unit_count, (int, float)):
                    continue

                current_rent = ws_prop.cell(row=386, column=current_col).value
                if current_rent and isinstance(current_rent, (int, float)):
                    current_weighted_sum += unit_count * current_rent
                    current_total_units += unit_count

            if current_total_units > 0:
                weighted_average_current = current_weighted_sum / current_total_units
                logger.info(f"  Current month weighted average rent: ${weighted_average_current:.2f}")

                # Calculate growth rate
                if weighted_average_prior > 0:
                    growth_rate = (weighted_average_current - weighted_average_prior) / weighted_average_prior
                    logger.info(f"  Annualized rent growth (prior year): {growth_rate:.6f}")
                    return growth_rate

            logger.warning("Could not calculate current month growth - no valid data")
            return None
        else:
            logger.warning("Could not calculate prior year growth - no valid data")
            return None

    except Exception as e:
        logger.error(f"Error in SUMPRODUCT calculation: {e}")
        return None

def extract_financial_metrics():
    """Extract NOI and CapEx from latest Monthly Financial Reviews file"""
    try:
        logger.info(f"Searching for latest Monthly Financial Reviews file...")

        if not FINANCIAL_PATH.exists():
            logger.warning(f"Financial reviews folder not found: {FINANCIAL_PATH}")
            return {}

        files = list(FINANCIAL_PATH.glob("*.xlsx")) + list(FINANCIAL_PATH.glob("*.xlsm"))
        if not files:
            logger.warning(f"No Excel files found in {FINANCIAL_PATH}")
            return {}

        latest_file = max(files, key=os.path.getctime)
        logger.info(f"Using latest file: {latest_file.name}")

        wb = load_workbook(latest_file, data_only=True)
        ws = wb['Summary']

        noi_value = extract_row_83_metric(ws, "NOI")
        capex_value = extract_row_83_metric(ws, "CapEx")

        if noi_value:
            logger.info(f"[OK] NOI Variance to Budget: {noi_value}")
        if capex_value:
            logger.info(f"[OK] CapEx vs Budget: {capex_value}")

        results = {}
        if noi_value:
            results['NOI Variance to Budget'] = noi_value
        if capex_value:
            results['CapEx vs Budget'] = capex_value

        return results

    except Exception as e:
        logger.error(f"Error extracting financial metrics: {e}")
        return {}

def extract_row_83_metric(ws, metric_type):
    """Extract NOI or CapEx metrics from row 83"""
    try:
        row_data = []

        for col_num in range(1, 100):
            cell = ws.cell(row=83, column=col_num)
            if cell.value and isinstance(cell.value, (int, float)):
                row_data.append(cell.value)

        if row_data:
            total = sum(row_data)
            return total
        else:
            return None

    except Exception as e:
        logger.warning(f"Could not extract {metric_type} from row 83: {e}")
        return None

def format_percentage(value):
    """Format a decimal value as percentage string"""
    if value is None:
        return "—"
    if isinstance(value, (int, float)):
        return f"{value*100:.2f}%"
    return str(value)

def save_metrics(metrics_dict):
    """Save metrics to local JSON cache"""
    try:
        data = {
            'timestamp': datetime.now().isoformat(),
            'metrics': [
                {
                    'name': key,
                    'value': value,
                    'updated': datetime.now().strftime('%Y-%m-%d'),
                    'source': get_source(key)
                }
                for key, value in metrics_dict.items()
            ],
            'count': len(metrics_dict),
            'success': True
        }

        with open(METRICS_FILE, 'w') as f:
            json.dump(data, f, indent=2)

        logger.info(f"[OK] Saved {len(metrics_dict)} metrics to {METRICS_FILE}")
        return True

    except Exception as e:
        logger.error(f"Error saving metrics: {e}")
        return False

def get_source(metric_name):
    """Return data source for each metric"""
    sources = {
        'Portfolio Occupancy': 'Occupancy Analysis.xlsx',
        'Leased Occupancy': 'Occupancy Analysis.xlsx',
        'Percent Collected': 'DQ Reports.xlsm',
        'Annualized Rent Growth': 'Effective Rent Analysis.xlsx',
        'Annualized Rent Growth (Prior Year)': 'Effective Rent Analysis.xlsx',
        'NOI Variance to Budget': 'Monthly Financial Reviews',
        'CapEx vs Budget': 'Monthly Financial Reviews'
    }
    return sources.get(metric_name, 'Unknown')

def main():
    try:
        logger.info("=" * 70)
        logger.info("Starting Enhanced Afton Metrics Extraction")
        logger.info("=" * 70)

        all_metrics = {}

        # Extract from all sources
        all_metrics.update(extract_occupancy_metrics())
        all_metrics.update(extract_dq_metrics())
        all_metrics.update(extract_effective_rent_metrics())
        all_metrics.update(extract_financial_metrics())

        # Save to cache
        if all_metrics:
            save_metrics(all_metrics)
            logger.info(f"[OK] Successfully extracted {len(all_metrics)} metrics")
        else:
            logger.warning("No metrics extracted")

        logger.info("=" * 70)
        logger.info("Metrics extraction completed")
        logger.info("=" * 70)

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)

if __name__ == '__main__':
    main()
