#!/usr/bin/env python3
"""
Extract individual reports from Scheduler_Reports.xlsx files
These files contain multiple worksheets - each one needs to be extracted as a separate file
"""

import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"scheduler_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")

def sanitize_filename(name):
    """Remove/replace invalid filename characters"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()

def extract_scheduler_reports():
    """
    Find all Scheduler_Reports*.xlsx files and extract their worksheets
    """
    logger.info("=" * 70)
    logger.info("SCHEDULER REPORTS EXTRACTOR")
    logger.info("=" * 70)
    logger.info(f"Looking for Scheduler_Reports files in: {DASHBOARD_DATA_DIR}\n")

    # Find all Scheduler_Reports files
    scheduler_files = list(DASHBOARD_DATA_DIR.glob("Scheduler_Reports*.xlsx"))

    if not scheduler_files:
        logger.warning("No Scheduler_Reports files found")
        return

    logger.info(f"Found {len(scheduler_files)} Scheduler_Reports files:\n")

    total_extracted = 0
    files_to_delete = []

    for scheduler_file in sorted(scheduler_files):
        logger.info(f"[Processing] {scheduler_file.name}")

        try:
            wb = load_workbook(scheduler_file)
            logger.info(f"  Worksheets: {len(wb.sheetnames)}")
            logger.info(f"  Names: {', '.join(wb.sheetnames)}\n")

            # Extract each worksheet
            for sheet_name in wb.sheetnames:
                try:
                    logger.info(f"  Extracting: '{sheet_name}'")

                    # Create new workbook with just this sheet
                    new_wb = load_workbook(scheduler_file)

                    # Keep only this sheet, delete others
                    sheets_to_delete = [s for s in new_wb.sheetnames if s != sheet_name]
                    for sheet_to_delete in sheets_to_delete:
                        del new_wb[sheet_to_delete]

                    # Rename the kept sheet
                    new_ws = new_wb.active
                    new_ws.title = "Report"

                    # Create output filename
                    safe_sheet_name = sanitize_filename(sheet_name)
                    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                    output_filename = f"{safe_sheet_name}_{timestamp}.xlsx"
                    output_path = DASHBOARD_DATA_DIR / output_filename

                    # Save the individual report
                    new_wb.save(output_path)
                    logger.info(f"    ✓ Saved: {output_filename}")
                    total_extracted += 1

                except Exception as e:
                    logger.error(f"    ✗ Error extracting '{sheet_name}': {e}")
                    continue

            # Mark this file for deletion
            files_to_delete.append(scheduler_file)

        except Exception as e:
            logger.error(f"  Error processing {scheduler_file.name}: {e}")
            continue

    # Delete original Scheduler_Reports files
    logger.info("\n" + "=" * 70)
    logger.info("CLEANING UP")
    logger.info("=" * 70)

    for file_to_delete in files_to_delete:
        try:
            logger.info(f"  Deleting: {file_to_delete.name}")
            file_to_delete.unlink()
        except Exception as e:
            logger.error(f"  Error deleting {file_to_delete.name}: {e}")

    # Summary
    logger.info("\n" + "=" * 70)
    logger.info("EXTRACTION COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Scheduler_Reports files processed: {len(scheduler_files)}")
    logger.info(f"Individual reports extracted: {total_extracted}")
    logger.info(f"Original files deleted: {len(files_to_delete)}")
    logger.info("\nAll Scheduler_Reports worksheets are now individual files")
    logger.info("=" * 70)

if __name__ == '__main__':
    extract_scheduler_reports()
