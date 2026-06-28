#!/usr/bin/env python3
"""
Daily organization of downloaded reports
1. Extract zip files
2. Extract worksheets from multi-sheet Excel files
3. Read first row (A1) of each worksheet to get actual report name
4. Organize into folders by actual report name
"""

import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import zipfile
import shutil

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"organize_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")

def sanitize_foldername(name):
    """Remove/replace invalid folder name characters"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()

def extract_zips():
    """Extract all zip files to root"""
    zip_files = list(DASHBOARD_DATA_DIR.glob("*.zip"))

    for zip_file in zip_files:
        try:
            logger.info(f"[Extracting zip] {zip_file.name}")

            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                for file_info in zip_ref.filelist:
                    if file_info.filename.lower().endswith(('.xlsx', '.xls')):
                        extract_path = DASHBOARD_DATA_DIR / Path(file_info.filename).name

                        with zip_ref.open(file_info) as source, open(extract_path, 'wb') as target:
                            target.write(source.read())

            # Delete zip after extraction
            zip_file.unlink()
            logger.info(f"  ✓ Extracted and deleted")

        except Exception as e:
            logger.error(f"  Error: {e}")

def organize_reports():
    """Organize all root-level Excel files into report folders"""
    logger.info("=" * 70)
    logger.info("DAILY REPORT ORGANIZATION")
    logger.info("=" * 70)

    # First extract any zips
    zip_files = list(DASHBOARD_DATA_DIR.glob("*.zip"))
    if zip_files:
        logger.info(f"\n[ZIP EXTRACTION]")
        extract_zips()
        logger.info()

    # Find all Excel files in root
    root_files = [f for f in DASHBOARD_DATA_DIR.glob("*.xlsx") if f.is_file()]
    root_files += [f for f in DASHBOARD_DATA_DIR.glob("*.xls") if f.is_file()]

    if not root_files:
        logger.info("No files to organize in root folder")
        return

    logger.info(f"[ORGANIZATION]\nFound {len(root_files)} files to organize\n")

    organized_count = 0
    report_folders = {}

    for filepath in sorted(root_files):
        try:
            logger.info(f"[Processing] {filepath.name}")

            wb = load_workbook(filepath)
            logger.info(f"  Worksheets: {len(wb.sheetnames)}")

            # Process each worksheet
            for sheet_name in wb.sheetnames:
                try:
                    # Create new workbook with just this sheet
                    new_wb = load_workbook(filepath)

                    # Keep only this sheet
                    sheets_to_delete = [s for s in new_wb.sheetnames if s != sheet_name]
                    for sheet_to_delete in sheets_to_delete:
                        del new_wb[sheet_to_delete]

                    # Rename sheet
                    new_ws = new_wb.active
                    new_ws.title = "Report"

                    # Read first row (A1) for actual report name
                    report_name = new_ws['A1'].value

                    if not report_name:
                        logger.warning(f"    No name in A1, using worksheet: {sheet_name}")
                        report_name = sheet_name

                    report_name = str(report_name).strip()
                    logger.info(f"    Report: {report_name}")

                    # Create folder for this report type
                    safe_folder_name = sanitize_foldername(report_name)
                    report_folder = DASHBOARD_DATA_DIR / safe_folder_name
                    report_folder.mkdir(exist_ok=True)

                    if safe_folder_name not in report_folders:
                        report_folders[safe_folder_name] = 0

                    # Create output filename with timestamp
                    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                    output_filename = f"{safe_folder_name}_{timestamp}.xlsx"
                    output_path = report_folder / output_filename

                    # Save the individual report
                    new_wb.save(output_path)
                    logger.info(f"    ✓ {safe_folder_name}/{output_filename}")

                    report_folders[safe_folder_name] += 1
                    organized_count += 1

                except Exception as e:
                    logger.error(f"    Error: {e}")

            # Delete original after processing all sheets
            filepath.unlink()
            logger.info(f"  ✓ Deleted original\n")

        except Exception as e:
            logger.error(f"Error: {e}\n")

    # Summary
    logger.info("=" * 70)
    logger.info("ORGANIZATION COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Files processed: {len(root_files)}")
    logger.info(f"Worksheets organized: {organized_count}")
    logger.info(f"\nReports organized into folders:")
    for folder_name in sorted(report_folders.keys()):
        logger.info(f"  - {folder_name}/ ({report_folders[folder_name]} files)")
    logger.info("=" * 70)

if __name__ == '__main__':
    organize_reports()
