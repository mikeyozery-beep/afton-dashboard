#!/usr/bin/env python3
"""
Smart YARDI Report Downloader
- Downloads all existing reports (first run)
- Downloads only new reports since last run (subsequent runs)
- Avoids duplicates by tracking timestamps
"""

import os
import re
import logging
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

try:
    import win32com.client
except ImportError:
    print("ERROR: pywin32 not installed")
    print("Run: pip install pywin32 --break-system-packages")
    exit(1)

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"yardi_smart_download_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
OUTPUT_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")
OUTLOOK_INBOX = "mikereports@aftonprop.com"
SENDER_EMAIL = "cdr@yardi.com"
TEMP_DIR = SCRIPT_DIR / "temp_yardi_downloads"
LAST_RUN_FILE = SCRIPT_DIR / ".yardi_last_run.json"

# Create directories
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TEMP_DIR.mkdir(parents=True, exist_ok=True)

def get_last_run_time():
    """Get the timestamp of the last successful run"""
    if LAST_RUN_FILE.exists():
        try:
            with open(LAST_RUN_FILE, 'r') as f:
                data = json.load(f)
                # Convert ISO format back to datetime
                return datetime.fromisoformat(data['last_run'])
        except Exception as e:
            logger.warning(f"Could not read last run time: {e}")
            return None
    return None

def save_last_run_time():
    """Save the current time as the last successful run"""
    try:
        data = {
            'last_run': datetime.now().isoformat(),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        with open(LAST_RUN_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        logger.info(f"[OK] Saved last run timestamp")
    except Exception as e:
        logger.warning(f"Could not save last run time: {e}")

def get_outlook_inbox():
    """Connect to Outlook and get the specified inbox"""
    logger.info(f"Connecting to Outlook for {OUTLOOK_INBOX}...")
    try:
        try:
            outlook = win32com.client.GetObject(None, "Outlook.Application")
        except:
            outlook = win32com.client.Dispatch("Outlook.Application")

        mapi = outlook.GetNamespace("MAPI")

        logger.info("Searching for mailbox folders...")

        # Try to find the specific email account
        inbox = None

        # Method 1: Search through all folders
        for folder in mapi.Folders:
            folder_name = folder.Name.lower()
            logger.info(f"  Found folder: {folder.Name}")

            # Check if this folder matches our email
            if OUTLOOK_INBOX.lower() in folder_name or "mikereports" in folder_name:
                try:
                    inbox = folder.Folders("Inbox")
                    logger.info(f"[OK] Connected to inbox: {folder.Name}")
                    return inbox
                except:
                    continue

        # Method 2: Try to get inbox by email address
        if not inbox:
            logger.info("Trying to access by email address...")
            try:
                inbox = mapi.Folders(OUTLOOK_INBOX).Folders("Inbox")
                logger.info(f"[OK] Connected to {OUTLOOK_INBOX} inbox")
                return inbox
            except:
                pass

        # Method 3: Fallback to default inbox
        if not inbox:
            logger.warning("Could not find specific inbox, using default Inbox")
            inbox = mapi.GetDefaultFolder(6)  # 6 = olFolderInbox
            logger.info("[OK] Using default Inbox")
            return inbox

    except Exception as e:
        logger.error(f"Error connecting to Outlook: {e}")
        raise

def get_excel_title(filepath):
    """Extract the title/name from Excel workbook"""
    try:
        wb = load_workbook(filepath, data_only=True)

        if wb.properties.title:
            title = wb.properties.title.strip()
            return title

        if wb.sheetnames:
            title = wb.sheetnames[0]
            return title

        title = Path(filepath).stem
        return title

    except Exception as e:
        logger.error(f"  Error reading Excel file: {e}")
        return None

def clean_filename(name):
    """Remove invalid characters from filename"""
    invalid_chars = r'[<>:"/\\|?*]'
    cleaned = re.sub(invalid_chars, '', name)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned

def download_and_process_attachments(download_all=False):
    """
    Download YARDI report attachments

    Args:
        download_all (bool): If True, download ALL reports. If False, only download since last run.
    """
    logger.info("=" * 70)
    logger.info("YARDI SMART REPORT DOWNLOADER")
    logger.info("=" * 70)

    last_run = get_last_run_time()

    if download_all:
        logger.info("\n[MODE] Download ALL existing reports")
        cutoff_time = None
    else:
        if last_run:
            logger.info(f"\n[MODE] Download NEW reports since last run")
            logger.info(f"Last run: {last_run.strftime('%Y-%m-%d %H:%M:%S')}")
            cutoff_time = last_run
        else:
            logger.info("\n[MODE] First run - downloading ALL reports")
            cutoff_time = None

    try:
        inbox = get_outlook_inbox()

        logger.info(f"\nSearching for emails from {SENDER_EMAIL}...")
        downloaded_count = 0
        skipped_count = 0
        error_count = 0

        # Get all emails in inbox
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)  # Sort by received time, descending

        logger.info("Processing emails...")

        for i, item in enumerate(items):
            try:
                # Check if email is from YARDI
                if SENDER_EMAIL.lower() not in item.SenderEmailAddress.lower():
                    continue

                # Check if has attachments
                if item.Attachments.Count == 0:
                    continue

                email_date = item.ReceivedTime

                # Skip if before cutoff time
                if cutoff_time and email_date < cutoff_time:
                    logger.info(f"\n[Email {i+1}] Received: {email_date.strftime('%Y-%m-%d %H:%M:%S')} - SKIPPED (before last run)")
                    skipped_count += 1
                    continue

                date_str = email_date.strftime("%m.%d.%Y")

                logger.info(f"\n[Email {i+1}] Received: {email_date.strftime('%Y-%m-%d %H:%M:%S')}")
                logger.info(f"  From: {item.SenderEmailAddress}")
                logger.info(f"  Subject: {item.Subject}")
                logger.info(f"  Attachments: {item.Attachments.Count}")

                # Process each attachment
                for attachment in item.Attachments:
                    try:
                        filename = attachment.Filename
                        logger.info(f"\n  Processing: {filename}")

                        # Skip if not Excel file
                        if not filename.lower().endswith(('.xlsx', '.xls')):
                            logger.info(f"    Skipping (not Excel)")
                            continue

                        # Save to temp location
                        temp_path = TEMP_DIR / filename
                        attachment.SaveAsFile(str(temp_path))
                        logger.info(f"    Downloaded to temp")

                        # Read Excel title
                        excel_title = get_excel_title(str(temp_path))

                        if not excel_title:
                            logger.warning(f"    Could not determine title, skipping")
                            temp_path.unlink()
                            error_count += 1
                            continue

                        # Clean the title
                        clean_title = clean_filename(excel_title)

                        # Create final filename: title + date
                        final_filename = f"{clean_title} {date_str}.xlsx"
                        final_path = OUTPUT_DIR / final_filename

                        # If file exists, append number
                        if final_path.exists():
                            base = final_filename.replace('.xlsx', '')
                            counter = 1
                            while (OUTPUT_DIR / f"{base}_{counter}.xlsx").exists():
                                counter += 1
                            final_filename = f"{base}_{counter}.xlsx"
                            final_path = OUTPUT_DIR / final_filename

                        # Move from temp to final location
                        temp_path.rename(final_path)
                        logger.info(f"    ✓ Saved: {final_filename}")
                        downloaded_count += 1

                    except Exception as e:
                        logger.error(f"    Error processing attachment: {e}")
                        error_count += 1
                        continue

            except Exception as e:
                logger.error(f"  Error processing email: {e}")
                continue

        logger.info("\n" + "=" * 70)
        logger.info(f"DOWNLOAD COMPLETE")
        logger.info("=" * 70)
        logger.info(f"Downloaded: {downloaded_count} reports")
        logger.info(f"Skipped: {skipped_count} reports (before last run)")
        logger.info(f"Errors: {error_count}")
        logger.info(f"Saved to: {OUTPUT_DIR}")
        logger.info("=" * 70)

        # Save the run timestamp for next time
        if downloaded_count > 0 or error_count == 0:
            save_last_run_time()

        return downloaded_count, error_count

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        return 0, 1

def cleanup_temp():
    """Clean up temporary files"""
    try:
        for file in TEMP_DIR.glob("*"):
            file.unlink()
        TEMP_DIR.rmdir()
        logger.info("Cleaned up temporary files")
    except Exception as e:
        logger.warning(f"Could not clean temp directory: {e}")

if __name__ == '__main__':
    import sys

    # Check command line arguments
    download_all = "--all" in sys.argv

    try:
        downloaded, errors = download_and_process_attachments(download_all=download_all)
        cleanup_temp()

        if errors > 0:
            exit(1)
    except KeyboardInterrupt:
        logger.warning("Interrupted by user")
        cleanup_temp()
        exit(1)
