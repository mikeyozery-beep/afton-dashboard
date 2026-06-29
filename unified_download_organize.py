#!/usr/bin/env python3
"""
UNIFIED YARDI DOWNLOAD & ORGANIZATION
Single-step automation: Download → Extract Zips → Extract Worksheets → Organize by Report Name
No manual sequencing required. One command handles everything.
"""

import logging
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
import zipfile
import shutil
import subprocess
import sys

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"unified_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")

# Files that live in the Dashboard Data ROOT and must NOT be filed into report
# folders - build_metrics.py reads them from the root by exact name.
ORGANIZE_EXCLUDE = {"open positions.xlsx"}

def sanitize_foldername(name):
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()

# A1 is normally the report title (e.g. "Rent Roll", "rs_sql_..."). But some
# YARDI financial exports put the OWNER/ENTITY group on row 1 -
# e.g. "Afton Garden Style Properties (.ap-gs)" - and the real report title on
# row 2 ("Lifetime Prefund", "Budget Comparison- current"). The entity header is
# recognizable by its parenthetical entity code "(.xxx)"; no real report title
# contains that. When we see it, the report name comes from row 2.
_ENTITY_HEADER = re.compile(r'\(\s*\.\w')

def report_name_from(a1, a2, fallback):
    """Pick the report name from row 1, falling back to row 2 when row 1 is an
    entity/owner header rather than a report title."""
    a1s = str(a1).strip() if a1 not in (None, "") else ""
    a2s = str(a2).strip() if a2 not in (None, "") else ""
    if a1s and _ENTITY_HEADER.search(a1s) and a2s:
        return a2s            # row 1 is the entity group; row 2 is the report
    return a1s or fallback

def variant_suffix(report_name, header_rows):
    """Some YARDI reports share one title (row 1) but ship in different
    groupings/filters that must NOT share a folder. Derive a folder-name suffix
    from the column-header rows so each variant files separately.

    `header_rows` is a list of the first ~6 rows, each a list of non-empty,
    stripped cell strings. Returns "" when the report has no known variants.

    Confirmed variants (see the mappings memory / audit):
      - Aged Receivables: resident-level detail ("by Tenant", has a Resident
        column) vs property summary ("by Property").
      - Rent Roll: unit/resident detail ("by Unit") vs unit-type summary
        ("by Unit Type", whose header row starts "Property | Unit Type").
      - rs_sql_tr_payments_muvan: a legacy column name "receipt_reference"
        (vs the current "payment_reference") marks an older report config.
    """
    name = str(report_name).strip().lower()
    cells = {c.strip().lower() for row in header_rows for c in row}

    if name == "aged receivables":
        return "by Tenant" if "resident" in cells else "by Property"

    if name == "rent roll":
        for row in header_rows:
            low = [c.strip().lower() for c in row]
            if low[:2] == ["property", "unit type"]:
                return "by Unit Type"      # summary grouped by unit type
            if low[:1] == ["unit"]:
                return "by Unit"           # resident/unit-level detail
        return ""

    if name == "rs_sql_tr_payments_muvan":
        return "(receipt_reference)" if "receipt_reference" in cells else ""

    return ""

def folder_name_for(report_name, header_rows):
    """Full destination folder name = report name + any variant suffix."""
    suffix = variant_suffix(report_name, header_rows)
    return sanitize_foldername(f"{report_name} {suffix}".strip() if suffix else report_name)

def step_1_download_files():
    """Download all files from mikereports@aftonprop.com inbox"""
    logger.info("\n" + "=" * 70)
    logger.info("STEP 1: DOWNLOADING FILES FROM INBOX")
    logger.info("=" * 70)

    try:
        download_script = SCRIPT_DIR / "download_yardi_outlook_com.ps1"
        if not download_script.exists():
            logger.error(f"Download script not found: {download_script}")
            return False

        logger.info(f"Running: {download_script.name}")
        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(download_script)],
            capture_output=True,
            text=True,
            timeout=600
        )

        if result.returncode != 0:
            logger.error(f"Download failed:\n{result.stderr}")
            return False

        logger.info("✓ Download complete")
        return True

    except Exception as e:
        logger.error(f"Step 1 error: {e}", exc_info=True)
        return False

def step_2_extract_zips():
    """Extract all zip files in Dashboard Data"""
    logger.info("\n" + "=" * 70)
    logger.info("STEP 2: EXTRACTING ZIP FILES")
    logger.info("=" * 70)

    try:
        zip_files = list(DASHBOARD_DATA_DIR.glob("*.zip"))

        if not zip_files:
            logger.info("No zip files found")
            return True

        logger.info(f"Found {len(zip_files)} zip files")

        for zip_file in zip_files:
            try:
                logger.info(f"Extracting: {zip_file.name}")

                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                    for file_info in zip_ref.filelist:
                        if file_info.filename.lower().endswith(('.xlsx', '.xls')):
                            extract_path = DASHBOARD_DATA_DIR / Path(file_info.filename).name

                            with zip_ref.open(file_info) as source, open(extract_path, 'wb') as target:
                                target.write(source.read())

                            logger.info(f"  ✓ {Path(file_info.filename).name}")

                zip_file.unlink()
                logger.info(f"  ✓ Deleted zip file")

            except Exception as e:
                logger.error(f"Error extracting {zip_file.name}: {e}")

        logger.info("✓ Zip extraction complete")
        return True

    except Exception as e:
        logger.error(f"Step 2 error: {e}", exc_info=True)
        return False

def _unique_output_path(folder, base):
    """Collision-safe path: base_<timestamp>[_n].xlsx (never overwrites)."""
    ts = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    candidate = folder / f"{base}_{ts}.xlsx"
    n = 1
    while candidate.exists():
        candidate = folder / f"{base}_{ts}_{n}.xlsx"
        n += 1
    return candidate


def _read_header_rows(ws, nrows=6, ncols=10):
    """First `nrows` rows as lists of non-empty, stripped cell strings -
    enough to recover the report title (row 1/2) and the column-header rows
    that variant_suffix() keys on."""
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=nrows, max_col=ncols, values_only=True):
        rows.append([str(c).strip() for c in r if c not in (None, "")])
    return rows


def organize_data_dir(data_dir):
    """File every report into a folder named by its report title, plus a
    variant suffix (see variant_suffix) so reports that share a title but differ
    in grouping/filters - e.g. Aged Receivables by Tenant vs by Property - never
    share a folder.

    Single-sheet workbooks (the common YARDI case) are MOVED as-is - no
    re-save - which is the fast path. Multi-sheet workbooks are loaded once
    and split into one values-only file per sheet. Downstream extractors read
    wb.active, so the single sheet does not need renaming.
    """
    data_dir = Path(data_dir)
    files = sorted(
        f for f in (list(data_dir.glob("*.xlsx")) + list(data_dir.glob("*.xls")))
        if f.is_file() and f.name.lower() not in ORGANIZE_EXCLUDE
    )
    if not files:
        logger.info("No files to organize")
        return 0, 0

    logger.info(f"Organizing {len(files)} file(s)")
    organized = 0
    folders = set()

    for filepath in files:
        try:
            # Cheap lazy pass: sheet count + A1 of the first sheet
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheetnames = wb.sheetnames

            if len(sheetnames) == 1:
                ws = wb[sheetnames[0]]
                hdr = _read_header_rows(ws)
                a1 = hdr[0][0] if hdr and hdr[0] else None
                a2 = hdr[1][0] if len(hdr) > 1 and hdr[1] else None
                report_name = report_name_from(a1, a2, sheetnames[0])
                wb.close()

                folder = data_dir / folder_name_for(report_name, hdr)
                folder.mkdir(exist_ok=True)
                dest = _unique_output_path(folder, folder.name)
                shutil.move(str(filepath), str(dest))          # fast: no re-save
                folders.add(folder.name)
                organized += 1
                logger.info(f"  [move]  {filepath.name} -> {folder.name}/")
            else:
                wb.close()
                src = load_workbook(filepath, data_only=True)  # single full load
                for sheet_name in sheetnames:
                    src_ws = src[sheet_name]
                    hdr = _read_header_rows(src_ws)
                    a1 = hdr[0][0] if hdr and hdr[0] else None
                    a2 = hdr[1][0] if len(hdr) > 1 and hdr[1] else None
                    report_name = report_name_from(a1, a2, sheet_name)
                    new_wb = Workbook()
                    new_ws = new_wb.active
                    for row in src_ws.iter_rows(values_only=True):
                        new_ws.append(row)
                    folder = data_dir / folder_name_for(report_name, hdr)
                    folder.mkdir(exist_ok=True)
                    dest = _unique_output_path(folder, folder.name)
                    new_wb.save(dest)
                    new_wb.close()
                    folders.add(folder.name)
                    organized += 1
                    logger.info(f"  [split] {filepath.name} [{sheet_name}] -> {folder.name}/")
                src.close()
                filepath.unlink()

        except Exception as e:
            logger.error(f"  Error organizing {filepath.name}: {e}")

    logger.info(f"Organized {organized} report(s) into {len(folders)} folder(s)")
    return organized, len(folders)


def step_3_extract_worksheets():
    """Organize downloaded reports into per-report folders."""
    logger.info("\n" + "=" * 70)
    logger.info("STEP 3: EXTRACTING & ORGANIZING BY REPORT NAME")
    logger.info("=" * 70)
    try:
        organize_data_dir(DASHBOARD_DATA_DIR)
        return True
    except Exception as e:
        logger.error(f"Step 3 error: {e}", exc_info=True)
        return False

def step_4_mark_read():
    """Mark the downloaded emails as read - ONLY after organize succeeded."""
    logger.info("\n" + "=" * 70)
    logger.info("STEP 4: MARKING PROCESSED EMAILS AS READ")
    logger.info("=" * 70)

    try:
        mark_script = SCRIPT_DIR / "mark_read_pending.ps1"
        if not mark_script.exists():
            logger.error(f"Mark-read script not found: {mark_script}")
            return False

        result = subprocess.run(
            ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(mark_script)],
            capture_output=True,
            text=True,
            timeout=300
        )
        logger.info(result.stdout.strip())
        if result.returncode != 0:
            logger.error(f"Mark-read failed:\n{result.stderr}")
            return False

        logger.info("✓ Mark-read complete")
        return True

    except Exception as e:
        logger.error(f"Step 4 error: {e}", exc_info=True)
        return False

def main():
    logger.info("\n" + "=" * 70)
    logger.info("UNIFIED YARDI DOWNLOAD & ORGANIZATION")
    logger.info("=" * 70)
    logger.info(f"Directory: {DASHBOARD_DATA_DIR}\n")

    # Execute all steps
    if not step_1_download_files():
        logger.error("\n✗ Process stopped at download")
        return False

    if not step_2_extract_zips():
        logger.error("\n✗ Process stopped at zip extraction")
        return False

    if not step_3_extract_worksheets():
        logger.error("\n✗ Process stopped at worksheet extraction")
        return False

    # Only mark emails read once their attachments are safely organized into folders.
    # Downloads + organize are already durable (files filed, emails recorded in the
    # processed ledger), so a mark-read failure is NOT a data-loss event - emails
    # simply stay UNREAD. But we still return failure so the scheduled task reports
    # RED and the unread state gets noticed and retried next run.
    mark_ok = step_4_mark_read()
    if not mark_ok:
        logger.warning("\n⚠ Organize succeeded but marking emails read failed - "
                       "emails stay UNREAD (already filed; ledger prevents re-filing). "
                       "Flagging the run as failed so it's noticed; retries next run.")

    logger.info("\n" + "=" * 70)
    if mark_ok:
        logger.info("✓ ALL COMPLETE - DOWNLOADED, ORGANIZED & MARKED READ")
    else:
        logger.info("⚠ COMPLETE WITH WARNINGS - DOWNLOADED & ORGANIZED; MARK-READ FAILED")
    logger.info("=" * 70 + "\n")
    return mark_ok

if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
