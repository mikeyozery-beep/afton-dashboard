#!/usr/bin/env python3
"""
Split YARDI Multi-Report Workbooks into Individual Reports
Reads Excel files with multiple worksheets and saves each as separate file
Named after the worksheet/report name
"""

import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"yardi_split_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")

def sanitize_filename(name):
    """Remove/replace invalid filename characters"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()

def split_workbook(excel_file):
    """
    Split a multi-sheet workbook into individual files
    Each worksheet becomes a separate Excel file named after the worksheet
    Files are organized into folders by report type
    """
    try:
        logger.info(f"\n[Processing] {excel_file.name}")

        wb = load_workbook(excel_file)
        logger.info(f"  Worksheets found: {len(wb.sheetnames)}")

        if len(wb.sheetnames) == 1:
            logger.info(f"  Single sheet workbook - skipping (already individual)")
            return 0

        split_count = 0

        # Extract each worksheet as separate file
        for sheet_name in wb.sheetnames:
            try:
                logger.info(f"  Extracting: {sheet_name}")

                # Read the worksheet
                ws = wb[sheet_name]

                # Create new workbook with just this sheet
                new_wb = load_workbook(excel_file)

                # Keep only this sheet, delete others
                sheets_to_delete = [s for s in new_wb.sheetnames if s != sheet_name]
                for sheet_to_delete in sheets_to_delete:
                    del new_wb[sheet_to_delete]

                # Rename the kept sheet to just the report name (clean)
                new_ws = new_wb.active
                new_ws.title = "Report"  # Use generic name to avoid Excel restrictions

                # Create output filename: report name + timestamp
                safe_sheet_name = sanitize_filename(sheet_name)
                timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                output_filename = f"{safe_sheet_name}_{timestamp}.xlsx"

                # Create subfolder for this report type
                report_folder = DASHBOARD_DATA_DIR / safe_sheet_name
                report_folder.mkdir(exist_ok=True)

                output_path = report_folder / output_filename

                # Save the individual report in its folder
                new_wb.save(output_path)
                logger.info(f"    ✓ Saved to {safe_sheet_name}/{output_filename}")
                split_count += 1

            except Exception as e:
                logger.error(f"    ✗ Error extracting '{sheet_name}': {e}")
                continue

        return split_count

    except Exception as e:
        logger.error(f"  Error processing {excel_file.name}: {e}")
        return 0

def delete_original_multisheet_files():
    """
    Delete original multi-sheet workbooks after splitting
    Only deletes .xlsx/.xls files in root Dashboard Data folder
    Keeps the organized report folders
    """
    logger.info("\n" + "=" * 70)
    logger.info("CLEANING UP ORIGINAL MULTI-SHEET FILES")
    logger.info("=" * 70)

    deleted_count = 0

    # Only process files directly in Dashboard Data (not in subfolders)
    for excel_file in DASHBOARD_DATA_DIR.glob("*.xlsx"):
        try:
            wb = load_workbook(excel_file)

            # If file has more than 1 sheet, it's a multi-sheet original - delete it
            if len(wb.sheetnames) > 1:
                logger.info(f"  Deleting: {excel_file.name} ({len(wb.sheetnames)} sheets)")
                excel_file.unlink()
                deleted_count += 1

        except Exception as e:
            logger.warning(f"  Could not check {excel_file.name}: {e}")

    # Also check .xls files
    for excel_file in DASHBOARD_DATA_DIR.glob("*.xls"):
        try:
            wb = load_workbook(excel_file)
            if len(wb.sheetnames) > 1:
                logger.info(f"  Deleting: {excel_file.name} ({len(wb.sheetnames)} sheets)")
                excel_file.unlink()
                deleted_count += 1
        except Exception as e:
            logger.warning(f"  Could not check {excel_file.name}: {e}")

    return deleted_count

def process_all_workbooks():
    """Main function: split all multi-sheet workbooks"""
    logger.info("=" * 70)
    logger.info("YARDI WORKBOOK SPLITTER")
    logger.info("=" * 70)
    logger.info(f"Dashboard Data folder: {DASHBOARD_DATA_DIR}\n")

    if not DASHBOARD_DATA_DIR.exists():
        logger.error(f"Dashboard Data folder not found: {DASHBOARD_DATA_DIR}")
        return

    # Find all Excel files
    excel_files = list(DASHBOARD_DATA_DIR.glob("*.xlsx")) + list(DASHBOARD_DATA_DIR.glob("*.xls"))

    if not excel_files:
        logger.warning("No Excel files found in Dashboard Data folder")
        return

    logger.info(f"Found {len(excel_files)} Excel files\n")

    total_split = 0

    # Process each file
    for excel_file in sorted(excel_files):
        split_count = split_workbook(excel_file)
        total_split += split_count

    # Clean up original multi-sheet files
    deleted_count = delete_original_multisheet_files()

    # Summary
    logger.info("\n" + "=" * 70)
    logger.info("SPLIT COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Workbooks processed: {len(excel_files)}")
    logger.info(f"Individual reports extracted: {total_split}")
    logger.info(f"Original multi-sheet files deleted: {deleted_count}")
    logger.info("")
    logger.info("Folder Structure Created:")
    logger.info("  Dashboard Data/")
    logger.info("    ├── Aged_Receivables/")
    logger.info("    │   ├── Aged_Receivables_2025-06-24.xlsx")
    logger.info("    │   ├── Aged_Receivables_2025-06-25.xlsx")
    logger.info("    │   └── ...")
    logger.info("    ├── Unit_Occupancy/")
    logger.info("    │   ├── Unit_Occupancy_2025-06-24.xlsx")
    logger.info("    │   ├── Unit_Occupancy_2025-06-25.xlsx")
    logger.info("    │   └── ...")
    logger.info("    └── [Other Report Types]/")
    logger.info("")
    logger.info("Result: Reports organized by type with dated versions")
    logger.info("=" * 70)

if __name__ == '__main__':
    process_all_workbooks()
