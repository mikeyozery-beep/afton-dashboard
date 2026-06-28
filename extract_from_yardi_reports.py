#!/usr/bin/env python3
"""
Extract Metrics from YARDI Reports
Reads all Excel files in Dashboard Data folder and extracts metrics
"""

import json
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"yardi_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")
DATA_DIR = SCRIPT_DIR / "data"
METRICS_FILE = DATA_DIR / "metrics.json"

DATA_DIR.mkdir(exist_ok=True)

def extract_from_unit_occupancy(filepath):
    """Extract occupancy metrics from Unit Occupancy report"""
    try:
        logger.info(f"  Extracting from Unit Occupancy: {filepath.name}")
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        metrics = {}
        properties = []

        # Structure: Col A = Property, Col D = Occupancy %, Col E = 7 Day Occupancy %
        # Start from row 7 (row 1-6 are headers/dates)

        for row_idx, row in enumerate(ws.iter_rows(min_row=7, max_row=36, values_only=True), start=7):
            prop_name = row[0]  # Column A
            occupancy_today = row[3]  # Column D (Occupancy %)
            occupancy_7day = row[4]  # Column E (7 Day Occupancy %)

            if prop_name and occupancy_today is not None:
                try:
                    # Convert percentage string to float
                    occ_value = float(str(occupancy_today).replace('%', ''))
                    properties.append({
                        'name': prop_name,
                        'occupancy_today': occ_value,
                        'occupancy_7day': float(str(occupancy_7day).replace('%', '')) if occupancy_7day else None
                    })
                    logger.info(f"    {prop_name}: {occupancy_today}")
                except (ValueError, TypeError):
                    continue

        # Calculate portfolio occupancy (average)
        if properties:
            avg_occupancy = sum(p['occupancy_today'] for p in properties) / len(properties)
            metrics['Portfolio Occupancy'] = round(avg_occupancy, 2)
            metrics['Properties'] = properties
            logger.info(f"    Portfolio Avg: {avg_occupancy:.2f}%")

        return metrics
    except Exception as e:
        logger.warning(f"  Error extracting from Unit Occupancy: {e}")
        return {}

def extract_from_income_statement(filepath):
    """Extract financial metrics from Income Statement"""
    try:
        logger.info(f"  Extracting from Income Statement: {filepath.name}")
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        metrics = {}

        # Look for NOI and other financial metrics
        # These locations will need to be verified against actual report structure
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if 'NOI' in cell.value.upper():
                        # Found NOI row, get values
                        logger.info(f"    Found NOI reference: {cell.value}")

        return metrics
    except Exception as e:
        logger.warning(f"  Error extracting from Income Statement: {e}")
        return {}

def extract_from_rent_roll(filepath):
    """Extract property-level metrics from Rent Roll"""
    try:
        logger.info(f"  Extracting from Rent Roll: {filepath.name}")
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        properties = {}

        # Extract by property
        # Typically: Property Name | Occupancy | Rent | Collections
        # Will need to verify actual structure

        return properties
    except Exception as e:
        logger.warning(f"  Error extracting from Rent Roll: {e}")
        return {}

def analyze_report_structure(filepath):
    """Analyze the structure of an Excel report"""
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        logger.info(f"\n  Report: {filepath.name}")
        logger.info(f"  Sheet: {ws.title}")
        logger.info(f"  Dimensions: {ws.dimensions}")

        # Show first 10 rows to understand structure
        logger.info("  First 10 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            logger.info(f"    Row {i+1}: {row[:5]}")  # Show first 5 columns

        return True
    except Exception as e:
        logger.error(f"  Error analyzing: {e}")
        return False

def process_all_reports():
    """Process all Excel files in Dashboard Data folder"""
    logger.info("=" * 70)
    logger.info("YARDI REPORTS METRICS EXTRACTOR")
    logger.info("=" * 70)

    if not DASHBOARD_DATA_DIR.exists():
        logger.error(f"Dashboard Data folder not found: {DASHBOARD_DATA_DIR}")
        return

    # Find all Excel files
    excel_files = list(DASHBOARD_DATA_DIR.glob("*.xlsx")) + list(DASHBOARD_DATA_DIR.glob("*.xls"))

    if not excel_files:
        logger.warning("No Excel files found in Dashboard Data folder")
        logger.info(f"Path: {DASHBOARD_DATA_DIR}")
        return

    logger.info(f"\nFound {len(excel_files)} Excel files:")
    for f in sorted(excel_files):
        logger.info(f"  - {f.name}")

    # Analyze each file to understand structure
    logger.info("\n" + "=" * 70)
    logger.info("ANALYZING REPORT STRUCTURES")
    logger.info("=" * 70)

    all_metrics = {}

    for filepath in sorted(excel_files):
        logger.info(f"\n[Processing] {filepath.name}")

        # Analyze structure
        analyze_report_structure(filepath)

        # Extract based on report type
        filename_lower = filepath.name.lower()

        if 'unit_occupancy' in filename_lower or 'occupancy' in filename_lower:
            metrics = extract_from_unit_occupancy(filepath)
            all_metrics.update(metrics)
        elif 'box_score' in filename_lower:
            metrics = extract_from_box_score(filepath)
            all_metrics.update(metrics)
        elif 'income_statement' in filename_lower:
            metrics = extract_from_income_statement(filepath)
            all_metrics.update(metrics)
        elif 'rent_roll' in filename_lower:
            metrics = extract_from_rent_roll(filepath)
            all_metrics.update(metrics)
        else:
            logger.info(f"  Unknown report type, analyzing structure only")

    # Save extracted metrics
    logger.info("\n" + "=" * 70)
    logger.info("SAVING EXTRACTED METRICS")
    logger.info("=" * 70)

    if all_metrics:
        data = {
            'timestamp': datetime.now().isoformat(),
            'metrics': [
                {
                    'name': key,
                    'value': value,
                    'updated': datetime.now().strftime('%Y-%m-%d'),
                    'source': 'YARDI Reports'
                }
                for key, value in all_metrics.items()
            ],
            'count': len(all_metrics),
            'success': True
        }

        with open(METRICS_FILE, 'w') as f:
            json.dump(data, f, indent=2)

        logger.info(f"[OK] Saved {len(all_metrics)} metrics to {METRICS_FILE}")
    else:
        logger.warning("No metrics extracted")

    logger.info("\n" + "=" * 70)
    logger.info("EXTRACTION COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Processed: {len(excel_files)} reports")
    logger.info(f"Extracted: {len(all_metrics)} metrics")
    logger.info(f"Next: Map Excel columns to metrics and implement extraction logic")
    logger.info("=" * 70)

if __name__ == '__main__':
    process_all_reports()
