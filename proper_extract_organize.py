#!/usr/bin/env python3
"""
Proper report extraction and organization
1. Extract zip files
2. Extract worksheets from multi-sheet Excel files
3. Read first row (A1) of each worksheet to get actual report name
4. Organize into folders by actual report name
5. Clean up originals
"""

import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import zipfile
import shutil

# Setup logging
SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"proper_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
DASHBOARD_DATA_DIR = Path(r"C:\Users\MichaelOzery\OneDrive - Afton Properties\Old Dropbox\My PC (DESKTOP-5D77V89)\Mike Ozery\Dashboard Data")

def sanitize_foldername(name):
    """Remove/replace invalid folder name characters"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()

def get_report_name_from_worksheet(filepath):
    """
    Read the first row (A1) of the active worksheet to get the actual report name
    Returns the value in A1, or None if can't read
    """
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        # Get value from A1
        cell_value = ws['A1'].value
        if cell_value:
            return str(cell_value).strip()
    except:
        pass
    return None

def extract_zip_files():
    """Extract all zip files in Dashboard Data root to temp location"""
    logger.info("=" * 70)
    logger.info("STEP 1: EXTRACTING ZIP FILES")
    logger.info("=" * 70)

    zip_files = list(DASHBOARD_DATA_DIR.glob("*.zip"))

    if not zip_files:
        logger.info("No zip files found")
        return []

    logger.info(f"Found {len(zip_files)} zip files\n")

    extracted_files = []

    for zip_file in zip_files:
        try:
            logger.info(f"[Extracting] {zip_file.name}")

            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                for file_info in zip_ref.filelist:
                    if file_info.filename.lower().endswith(('.xlsx', '.xls')):
                        # Extract to root of Dashboard Data
                        extract_path = DASHBOARD_DATA_DIR / Path(file_info.filename).name

                        with zip_ref.open(file_info) as source, open(extract_path, 'wb') as target:
                            target.write(source.read())

                        logger.info(f"  ✓ Extracted: {Path(file_info.filename).name}")
                        extracted_files.append(extract_path)

        except Exception as e:
            logger.error(f"  Error extracting {zip_file.name}: {e}")

    logger.info(f"\nExtracted {len(extracted_files)} files from zips\n")
    return zip_files

def extract_and_organize_worksheets():
    """
    Extract worksheets from multi-sheet Excel files
    Read first row (A1) to get actual report name
    Organize into folders by report name
    """
    logger.info("=" * 70)
    logger.info("STEP 2: EXTRACTING WORKSHEETS & ORGANIZING BY NAME")
    logger.info("=" * 70)

    # Find all Excel files in root
    root_files = list(DASHBOARD_DATA_DIR.glob("*.xlsx")) + list(DASHBOARD_DATA_DIR.glob("*.xls"))

    if not root_files:
        logger.warning("No Excel files found in root")
        return [], {}

    logger.info(f"Found {len(root_files)} Excel files\n")

    files_to_delete = []
    organized_count = 0
    report_folders = {}

    for excel_file in sorted(root_files):
        try:
            wb = load_workbook(excel_file)
            logger.info(f"[Processing] {excel_file.name}")
            logger.info(f"  Worksheets: {len(wb.sheetnames)}")

            # Process each worksheet
            for sheet_name in wb.sheetnames:
                try:
                    logger.info(f"  Extracting: '{sheet_name}'")

                    # Create new workbook with just this sheet
                    new_wb = load_workbook(excel_file)

                    # Keep only this sheet
                    sheets_to_delete = [s for s in new_wb.sheetnames if s != sheet_name]
                    for sheet_to_delete in sheets_to_delete:
                        del new_wb[sheet_to_delete]

                    # Rename sheet
                    new_ws = new_wb.active
                    new_ws.title = "Report"

                    # Read first row (A1) to get actual report name
                    report_name = new_ws['A1'].value

                    if not report_name:
                        logger.warning(f"    ⚠ No report name in A1, using worksheet name: {sheet_name}")
                        report_name = sheet_name

                    report_name = str(report_name).strip()
                    logger.info(f"    Report Name: {report_name}")

                    # Create folder for this report type
                    safe_folder_name = sanitize_foldername(report_name)
                    report_folder = DASHBOARD_DATA_DIR / safe_folder_name
                    report_folder.mkdir(exist_ok=True)

                    if safe_folder_name not in report_folders:
                        report_folders[safe_folder_name] = 0

                    # Create output filename with timestamp
                    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                    output_filename = f"{safe_folder_name}_{timestamp}.xlsx"
                    output_path = report_folder / output_filename

                    # Save the individual report
                    new_wb.save(output_path)
                    logger.info(f"    ✓ Saved to {safe_folder_name}/{output_filename}")

                    report_folders[safe_folder_name] += 1
                    organized_count += 1

                except Exception as e:
                    logger.error(f"    Error extracting '{sheet_name}': {e}")

            # Mark for deletion after processing
            files_to_delete.append(excel_file)

        except Exception as e:
            logger.error(f"  Error processing {excel_file.name}: {e}")

    logger.info(f"\nOrganized {organized_count} worksheets into folders\n")
    return files_to_delete, report_folders

def cleanup_originals(excel_files_to_delete, zip_files_to_delete):
    """Delete original files after extraction"""
    logger.info("=" * 70)
    logger.info("STEP 3: CLEANING UP ORIGINALS")
    logger.info("=" * 70)

    deleted_count = 0

    # Delete Excel files
    for excel_file in excel_files_to_delete:
        try:
            logger.info(f"  Deleting: {excel_file.name}")
            excel_file.unlink()
            deleted_count += 1
        except Exception as e:
            logger.error(f"  Error deleting {excel_file.name}: {e}")

    # Delete zip files
    for zip_file in zip_files_to_delete:
        try:
            logger.info(f"  Deleting: {zip_file.name}")
            zip_file.unlink()
            deleted_count += 1
        except Exception as e:
            logger.error(f"  Error deleting {zip_file.name}: {e}")

    logger.info(f"\nDeleted {deleted_count} original files\n")

def cleanup_old_folders():
    """Delete old poorly-organized folders (Report, Report1, etc.)"""
    logger.info("=" * 70)
    logger.info("STEP 0: REMOVING OLD FOLDERS")
    logger.info("=" * 70)

    # List of old folder names to remove
    old_folders = ['Report', 'Report1', 'Report_', 'Report1_']

    for folder_name in old_folders:
        folder_path = DASHBOARD_DATA_DIR / folder_name
        if folder_path.exists() and folder_path.is_dir():
            try:
                logger.info(f"  Removing: {folder_name}/")
                shutil.rmtree(folder_path)
                logger.info(f"    ✓ Removed")
            except Exception as e:
                logger.error(f"  Error removing {folder_name}: {e}")

    logger.info("")

def main():
    """Main extraction and organization process"""
    logger.info("=" * 70)
    logger.info("PROPER REPORT EXTRACTION & ORGANIZATION")
    logger.info("=" * 70)
    logger.info(f"Directory: {DASHBOARD_DATA_DIR}\n")

    # Step 0: Remove old folders
    cleanup_old_folders()

    # Step 1: Extract zips
    zip_files = extract_zip_files()

    # Step 2: Extract worksheets and organize
    result = extract_and_organize_worksheets()
    excel_files = result[0] if result else []
    report_folders = result[1] if result else {}

    # Step 3: Cleanup originals
    cleanup_originals(excel_files, zip_files)

    # Summary
    logger.info("=" * 70)
    logger.info("COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Zip files extracted: {len(zip_files)}")
    logger.info(f"Excel files processed: {len(excel_files)}")
    logger.info(f"Worksheets extracted: {sum(report_folders.values())}")
    logger.info(f"\nReport Folders Created:")
    for folder_name in sorted(report_folders.keys()):
        logger.info(f"  - {folder_name}/ ({report_folders[folder_name]} files)")
    logger.info("\n✓ All reports properly organized by actual report name")
    logger.info("=" * 70)

if __name__ == '__main__':
    main()
